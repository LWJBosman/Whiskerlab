# -*- coding: utf-8 -*-
"""
Created on Tue Mar  7 10:39:37 2023

@author: Randy
"""
import pandas as pd 
import seaborn as sns
import matplotlib as plt
import matplotlib.pyplot as plt 
import numpy as np
from sklearn.decomposition import PCA
from shapely.geometry import Point

from sklearn.preprocessing import StandardScaler
#%% median trial times per mouse per session 
ercc1_mut = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\mutant\\Parameters.xlsx")
ercc1_ctr = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\control\\Parameters.xlsx")


df2 = pd.DataFrame(ercc1_mut)
dff2 = df2.groupby(["subject id","session nr"]).median()
dff2= dff2["timerun"]
df = pd.DataFrame(ercc1_ctr)
dff = df.groupby(["subject id","session nr"]).median()
dff= dff["timerun"]

test = [dff,dff2]
result = pd.concat(test, axis=1)
with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\median_runtime_per_session_per_mouse.xlsx", engine='xlsxwriter') as writer:
    result.to_excel(writer)
    
#%%median touchtime per mouser per session
ercc1_mut = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\mutant\\alltouches.xlsx")
ercc1_ctr = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\control\\alltouches.xlsx")

if (ercc1_mut["sides"] == 1).all():
    ercc1_mut=ercc1_mut.drop(ercc1_mut[ercc1_mut.rung%2!=0].index)
if (ercc1_mut["sides"] == 0).all():
    ercc1_mut=ercc1_mut.drop(ercc1_mut[ercc1_mut.rung%2==0].index)
if (ercc1_ctr["sides"] == 1).all():
    ercc1_ctr=ercc1_ctr.drop(ercc1_ctr[ercc1_ctr.rung%2!=0].index)
if (ercc1_ctr["sides"] == 0).all():
    ercc1_ctr=ercc1_ctr.drop(ercc1_ctr[ercc1_ctr.rung%2==0].index)

df2 = pd.DataFrame(ercc1_mut)
df2.pop("rung")
dff2 = df2.groupby(["subject id","sessionnr"]).median()
dff2= dff2["touch"]
df = pd.DataFrame(ercc1_ctr)
df.pop("rung")
dff = df.groupby(["subject id","sessionnr"]).median()
dff= dff["touch"]

test = [dff, dff2]
result = pd.concat(test, axis=1)
with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\median_touchtime_per_session_per_mouse.xlsx", engine='xlsxwriter') as writer:
    result.to_excel(writer)
    
#%% step types per muis per session
ercc1_mut = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\sorted\\mutant\\Steps_front.xlsx")
ercc1_ctr = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\control\\Steps_front.xlsx")


df = pd.DataFrame (ercc1_mut)
df2 = pd.DataFrame(ercc1_ctr)
dff2 = df2.groupby(["subject id","session nr"]).sum()
dff2.pop("session id")
dff2.pop("run id")
dff2.pop("direction")
dff = df.groupby(["subject id","session nr"]).sum()
dff.pop("session id")
dff.pop("run id")
dff.pop("direction")
print(dff)
with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\step_type_per_mouse_per_session_ctr.xlsx", engine='xlsxwriter') as writer:
    dff.to_excel(writer, sheet_name='MUT')
    dff2.to_excel(writer,sheet_name='CTR')
    
#%% all runtimes
y1=[]
y2=[]
y3=[]
y4=[]
y5=[]
y6=[]
ercc1_mut = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\mutant\\Parameters.xlsx")
ercc1_ctr = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\control\\Parameters.xlsx")

ercc1_mut = ercc1_mut.drop(ercc1_mut[ercc1_mut.timerun < 800].index)
ercc1_mut = ercc1_mut.drop(ercc1_mut[ercc1_mut.timerun > 15000].index)
ercc1_ctr = ercc1_ctr.drop(ercc1_ctr[ercc1_ctr.timerun > 15000].index)
ercc1_ctr = ercc1_ctr.drop(ercc1_ctr[ercc1_ctr.timerun < 800].index)

ko_6 = ercc1_mut[ercc1_mut["session nr"] == 6]
ko_12 = ercc1_mut[ercc1_mut["session nr"] == 12]
ko_18 = ercc1_mut[ercc1_mut["session nr"] == 18]
wt_6 = ercc1_ctr[ercc1_ctr["session nr"] == 6]
wt_12 = ercc1_ctr[ercc1_ctr["session nr"] == 12]
wt_18 = ercc1_ctr[ercc1_ctr["session nr"] == 18]

y1.extend(ko_6["timerun"])
y2.extend(ko_12["timerun"])
y3.extend(ko_18["timerun"])
y4.extend(wt_6["timerun"])
y5.extend(wt_12["timerun"])
y6.extend(wt_18["timerun"])

plt.figure()
bw=0.35

sns.distplot(y1,label=('knockout session 6'), hist = False, kde = True, kde_kws = {'bw': bw})
sns.distplot(y2,label=('knockout session 12'), hist = False, kde = True, kde_kws = {'bw': bw})
sns.distplot(y3,label=('knockout session 18'), hist = False, kde = True, kde_kws = {'bw': bw})
sns.distplot(y4,label=('wildtype session 6'), hist = False, kde = True, kde_kws = {'bw': bw})
sns.distplot(y5,label=('wildtype session 12'), hist = False, kde = True, kde_kws = {'bw': bw})
sns.distplot(y6,label=('wildtype session 18'), hist = False, kde = True, kde_kws = {'bw': bw})
plt.legend()
plt.xlabel('runtime')
plt.xlim(0,8000)
plt.savefig("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\runtimes.eps")

#%% all touchtimes
y1=[]
y2=[]
y3=[]
y4=[]
y5=[]
y6=[]
ercc1_mut = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\mutant\\alltouches.xlsx")
ercc1_ctr = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\control\\alltouches.xlsx")

ercc1_ctr = ercc1_ctr.drop(ercc1_ctr[ercc1_ctr.sessionnr >= 19].index)
ercc1_ctr = ercc1_ctr.drop(ercc1_ctr[ercc1_ctr.sessionnr < 6].index)


ko_6 = ercc1_mut[ercc1_mut["sessionnr"] == 6]
ko_12 = ercc1_mut[ercc1_mut["sessionnr"] == 12]
ko_18 = ercc1_mut[ercc1_mut["sessionnr"] == 18]
wt_6 = ercc1_ctr[ercc1_ctr["sessionnr"] == 6]
wt_12 = ercc1_ctr[ercc1_ctr["sessionnr"] == 12]
wt_18 = ercc1_ctr[ercc1_ctr["sessionnr"] == 18]

y1.extend(ercc1_ctr["touch"])
y2.extend(ko_12["touch"])
y3.extend(ko_18["touch"])
y4.extend(wt_6["touch"])
y5.extend(wt_12["touch"])
y6.extend(wt_18["touch"])


df1 = pd.DataFrame(y1)
df2 = pd.DataFrame(y2)
df3 = pd.DataFrame(y3)
df4 = pd.DataFrame(y4)
df5 = pd.DataFrame(y5)
df6 = pd.DataFrame(y6)
df_final =[df1,df2,df3,df4,df5,df6]
test= pd.concat(df_final,axis=1)
test.columns = ["ko_6", "ko_12", "ko_18","wt_6","wt_12","wt_18"]

cut_off_time = 500

y11 = [item for item in y1 if item <= cut_off_time]
y21 = [item for item in y2 if item <= cut_off_time]
y31 = [item for item in y3 if item <= cut_off_time]
y41 = [item for item in y4 if item <= cut_off_time]
y51 = [item for item in y5 if item <= cut_off_time]
y61 = [item for item in y6 if item <= cut_off_time]


with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\touchtime(rawdata)alltouches.xlsx", engine='xlsxwriter') as writer:
      df1.to_excel(writer, sheet_name = 'Touches')
bw=0.12
plt.figure()
sns.distplot(y11,label=('knockout session 8'),  hist = False, kde = True, kde_kws = {'bw' : bw})
sns.distplot(y21,label=('knockout session 12'), hist = False, kde = True, kde_kws = {'bw' : bw})
sns.distplot(y31,label=('knockout session 18'), hist = False, kde = True, kde_kws = {'bw' : bw})
sns.distplot(y41,label=('wildtype session 6'), hist = False, kde = True, kde_kws = {'bw' : bw})
sns.distplot(y51,label=('wildtype session 12'), hist = False, kde = True, kde_kws = {'bw' : bw})
sns.distplot(y61,label=('wildtype session 18'),hist = False, kde = True, kde_kws = {'bw' : bw})
plt.xlabel('touch time')
plt.ylabel('amount of touches')
plt.xlim(0,500)
plt.legend()
plt.savefig("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\touchtime all touches.eps")


#%% runtypes 
df = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\mutant\\Parameters.xlsx")
light=[]
air=[]
returns=[]
escape=[]
irregular=[]
irregularses=[]
airses=[]
escapeses=[]
returnsses=[]
lightses=[]

# Loop through the rows of the dataframe
for index, row in df.iterrows():
  if ('4-6-4-6' in row['state sequence']) & ('1-2-1' in row ["state sequence"]):
    irregular.append(row['subject id'])
    irregularses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '1-2-1' in row['state sequence']:
    escape.append(row['subject id'])
    escapeses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '4-6-4-6' in row['state sequence']:
    returns.append(row['subject id'])
    returnsses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '1-3-6' in row['state sequence']:
    light.append(row['subject id'])
    lightses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '1-3-4-6' in row['state sequence']:
    air.append(row['subject id'])
    airses.append(row['session nr'])
    df.drop(index, inplace=True)



escape = pd.DataFrame(escape)
returns = pd.DataFrame(returns)
light = pd.DataFrame(light)
air = pd.DataFrame(air)
irregular = pd.DataFrame(irregular)
escapeses = pd.DataFrame(escapeses)
airses = pd.DataFrame(airses)
lightses = pd.DataFrame(lightses)
returnsses = pd.DataFrame(returnsses)
irregularses = pd.DataFrame(irregularses)

df_final = [escape,escapeses]
df_final2= [returns,returnsses]
df_final3 = [air,airses]
df_final4 = [light,lightses]
df_final5 =[irregular,irregularses]
df_final6 = [escape,escapeses,returns,returnsses,air,airses,light,lightses,irregular,irregularses]
test=pd.concat(df_final,axis=1)
test2 = pd.concat(df_final2,axis=1)
test3 = pd.concat(df_final3,axis=1)
test4 = pd.concat(df_final4,axis=1)
test5 = pd.concat(df_final5,axis=1)
test6 = pd.concat(df_final6, axis=1)
test.columns = ["subjectid","session nr"]
test2.columns=["subjectid","session nr"]
test3.columns=["subjectid","session nr"]
test4.columns=["subjectid","session nr"]
test5.columns=["subjectid","session nr"]
test=test.groupby(["subjectid","session nr"]).size()
test2=test2.groupby(["subjectid","session nr"]).size()
test3=test3.groupby(["subjectid","session nr"]).size()
test4=test4.groupby(["subjectid","session nr"]).size()
test5=test5.groupby(["subjectid","session nr"]).size()

final_test = [test3,test2,test,test4,test5]
final_test=pd.concat(final_test,axis=1)
final_test.columns = ["air","return","escape","light","irregular"]
with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\runtypes_mut.xlsx", engine='xlsxwriter') as writer:
      final_test.to_excel(writer, sheet_name = 'all')
df = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\control\\Parameters.xlsx")
light=[]
air=[]
returns=[]
escape=[]
irregular=[]
irregularses=[]
airses=[]
escapeses=[]
returnsses=[]
lightses=[]

# Loop through the rows of the dataframe
for index, row in df.iterrows():
  if ('4-6-4-6' in row['state sequence']) & ('1-2-1' in row ["state sequence"]):
    irregular.append(row['subject id'])
    irregularses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '1-2-1' in row['state sequence']:
    escape.append(row['subject id'])
    escapeses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '4-6-4-6' in row['state sequence']:
    returns.append(row['subject id'])
    returnsses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '1-3-6' in row['state sequence']:
    light.append(row['subject id'])
    lightses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '1-3-4-6' in row['state sequence']:
    air.append(row['subject id'])
    airses.append(row['session nr'])
    df.drop(index, inplace=True)



escape = pd.DataFrame(escape)
returns = pd.DataFrame(returns)
light = pd.DataFrame(light)
air = pd.DataFrame(air)
irregular = pd.DataFrame(irregular)
escapeses = pd.DataFrame(escapeses)
airses = pd.DataFrame(airses)
lightses = pd.DataFrame(lightses)
returnsses = pd.DataFrame(returnsses)
irregularses = pd.DataFrame(irregularses)

df_final = [escape,escapeses]
df_final2= [returns,returnsses]
df_final3 = [air,airses]
df_final4 = [light,lightses]
df_final5 =[irregular,irregularses]
df_final6 = [escape,escapeses,returns,returnsses,air,airses,light,lightses,irregular,irregularses]
test=pd.concat(df_final,axis=1)
test2 = pd.concat(df_final2,axis=1)
test3 = pd.concat(df_final3,axis=1)
test4 = pd.concat(df_final4,axis=1)
test5 = pd.concat(df_final5,axis=1)
test6 = pd.concat(df_final6, axis=1)
test.columns = ["subjectid","session nr"]
test2.columns=["subjectid","session nr"]
test3.columns=["subjectid","session nr"]
test4.columns=["subjectid","session nr"]
test5.columns=["subjectid","session nr"]
test=test.groupby(["subjectid","session nr"]).size()
test2=test2.groupby(["subjectid","session nr"]).size()
test3=test3.groupby(["subjectid","session nr"]).size()
test4=test4.groupby(["subjectid","session nr"]).size()
test5=test5.groupby(["subjectid","session nr"]).size()

final_test = [test3,test2,test,test4,test5]
final_test=pd.concat(final_test,axis=1)
final_test.columns = ["air","return","escape","light","irregular"]
with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\runtypes_ctr.xlsx", engine='xlsxwriter') as writer:
      final_test.to_excel(writer, sheet_name = 'all')
#%% duty cycle / cv2  per muis per sessie 
swingtimeRF=[]
swingtime=[]
swingtimeRH=[]
swingtimeLH=[]

df = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\mutant\\alltouches.xlsx")
df = df.iloc[: , 1:]

dfLH = df
dfRF = df
dfRH = df

df = df[df["paw"].str.contains("rh") == False]
df = df[df["paw"].str.contains("lh") == False]
df = df[df["paw"].str.contains("rf") == False]

dfLH = dfLH[dfLH["paw"].str.contains("rh") == False]
dfLH = dfLH[dfLH["paw"].str.contains("lf") == False]
dfLH = dfLH[dfLH["paw"].str.contains("rf") == False]

dfRH = dfRH[dfRH["paw"].str.contains("lh") == False]
dfRH = dfRH[dfRH["paw"].str.contains("lf") == False]
dfRH = dfRH[dfRH["paw"].str.contains("rf") == False]

dfRF = dfRF[dfRF["paw"].str.contains("lh") == False]
dfRF = dfRF[dfRF["paw"].str.contains("lf") == False]
dfRF = dfRF[dfRF["paw"].str.contains("rh") == False]

df["diff"] = df["run id"].diff()
dfLH["diff"] = dfLH["run id"].diff()
dfRH["diff"] = dfRH["run id"].diff()
dfRF["diff"] = dfRF["run id"].diff()

    
with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\test.xlsx", engine='xlsxwriter') as writer:
      df.to_excel(writer, sheet_name = 'LF')
      dfLH.to_excel(writer, sheet_name = 'LH')
      dfRF.to_excel(writer, sheet_name = 'RF')
      dfRH.to_excel(writer, sheet_name = 'RH')


df2 = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\test.xlsx", sheet_name='LF', header = 0, index_col=None)
df2 = df2.iloc[: , 1:]
df2LH = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\test.xlsx", sheet_name='LH', header = 0, index_col=None)
df2LH = df2LH.iloc[: , 1:]
df2RH = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\test.xlsx", sheet_name='RH', header = 0, index_col=None)
df2RH = df2RH.iloc[: , 1:]
df2RF = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\test.xlsx", sheet_name='RF', header = 0, index_col=None)
df2RF = df2RF.iloc[: , 1:]

index_list = df2.index[df2["diff"] != 0].tolist()
index_listLH = df2LH.index[df2LH["diff"] != 0].tolist()
index_listRH = df2RH.index[df2RH["diff"] != 0].tolist()
index_listRF = df2RF.index[df2RF["diff"] != 0].tolist()


def subtract_from_columns(df2, index_list, value_column):
  for i in range(len(index_list) - 1):
    start_index = index_list[i]
    end_index = index_list[i+1]
    value = df2.at[start_index, value_column]
    df2.iloc[start_index:end_index, 8] -= value
    df2.iloc[start_index:end_index, 9] -= value
  return df2

value_column = 'begintouches'

result = subtract_from_columns(df2,index_list,value_column) 
for i in range(len(result)-1):
    if result["direction"].iloc[i] == 0:
        result["rung"][i] = (result["rung"][i] - 38)*-1
    swingtime.append(result["begintouches"].iloc[i+1]-result["endtouches"].iloc[i])
swingtime.insert(0,0)
result["swingtime"] = swingtime
mask = result['diff'] != 0
result.loc[mask, 'swingtime'] = pd.np.nan
            
resultLH = subtract_from_columns(df2LH,index_listLH,value_column) 
for i in range(len(resultLH)-1):
    if resultLH["direction"].iloc[i] == 0:
        resultLH["rung"][i] = (resultLH["rung"][i] - 38)*-1
    swingtimeLH.append(resultLH["begintouches"].iloc[i+1]-resultLH["endtouches"].iloc[i])
swingtimeLH.insert(0,0)   
resultLH["swingtime"] = swingtimeLH 
mask = resultLH['diff'] != 0
resultLH.loc[mask, 'swingtime'] = pd.np.nan
       
resultRH = subtract_from_columns(df2RH,index_listRH,value_column) 
for i in range(len(resultRH)-1):
    if resultRH["direction"].iloc[i] == 0:
        resultRH["rung"][i] = (resultRH["rung"][i] - 38)*-1
    swingtimeRH.append(resultRH["begintouches"].iloc[i+1]-resultRH["endtouches"].iloc[i])
swingtimeRH.insert(0,0)   
resultRH["swingtime"] = swingtimeRH
mask = resultRH['diff'] != 0
resultRH.loc[mask, 'swingtime'] = pd.np.nan
  
resultRF = subtract_from_columns(df2RF,index_listRF,value_column) 
for i in range(len(resultRF)-1):
    if resultRF["direction"].iloc[i] == 0:
        resultRF["rung"][i] = (resultRF["rung"][i] - 38)*-1
    swingtimeRF.append(resultRF["begintouches"].iloc[i+1]-resultRF["endtouches"].iloc[i])
swingtimeRF.insert(0,0)    
resultRF["swingtime"] = swingtimeRF
mask = resultRF['diff'] != 0
resultRF.loc[mask, 'swingtime'] = pd.np.nan


merge = [result,resultRF,resultRH,resultLH]
merger=pd.concat(merge,axis=0)
merger = merger.reset_index(drop=True)

indexes = []
indexes2= []
for index, row in merger.iterrows():
  if row['diff'] != 0:
    indexes2.append(index)
    
for index, row in merger.iterrows():
  if row['swingtime'] < 0:
    indexes.append(index)

# removeslist = find_range(indexes, indexes2)

# merger = merger.drop(removeslist)
# merger = merger.reset_index(drop=True)

with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\MUT.xlsx", engine='xlsxwriter') as writer:
      result.to_excel(writer, sheet_name = 'LF')
      resultLH.to_excel(writer, sheet_name = 'LH')
      resultRH.to_excel(writer, sheet_name = 'RH')
      resultRF.to_excel(writer, sheet_name = 'RF')
      merger.to_excel(writer, sheet_name = 'merger')

df3 = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\MUT.xlsx", sheet_name='merger', header = 0, index_col=None)
steptime = []
for i in range(len(df3)-1):
    steptime.append(df3["touch"].iloc[i]+df3["swingtime"].iloc[i+1])
steptime.insert(0,0)    
df3["steptime"]=steptime
df3 = df3.iloc[: , 1:]

df3.reset_index()


with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\MUT.xlsx", engine='xlsxwriter') as writer:
      result.to_excel(writer, sheet_name = 'LF')
      resultLH.to_excel(writer, sheet_name = 'LH')
      resultRH.to_excel(writer, sheet_name = 'RH')
      resultRF.to_excel(writer, sheet_name = 'RF')
      merger.to_excel(writer, sheet_name = 'merger')
      df3.to_excel(writer, sheet_name = 'merger+steptime')

df4 = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\MUT.xlsx", sheet_name= 'merger+steptime')
df2 = df4[['diff', 'swingtime','steptime']]
df4 = df4.drop(['diff', 'swingtime','steptime'], axis=1)

df2 =df2[1:]
df2 = df2.append(pd.Series([np.nan]*len(df2.columns), index=df2.columns), ignore_index=True)
df4 = df4.iloc[: , 1:]
df_concatMUT = pd.concat([df4, df2], axis=1)

df_concatMUT['DC'] = df_concatMUT['touch']/df_concatMUT['steptime']
cv2=[]
for i in range(len(df_concatMUT)-1):
    cv2.append(np.abs(2*(df_concatMUT["steptime"].iloc[i+1]-df_concatMUT["steptime"].iloc[i])/(df_concatMUT["steptime"].iloc[i+1]+df_concatMUT["steptime"].iloc[i])))
cv2.insert(-1,0)    
df_concatMUT["cv2"]=cv2

swingtimeRF=[]
swingtime=[]
swingtimeRH=[]
swingtimeLH=[]
df = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\control\\alltouches.xlsx")
df = df.iloc[: , 1:]

dfLH = df
dfRF = df
dfRH = df

df = df[df["paw"].str.contains("rh") == False]
df = df[df["paw"].str.contains("lh") == False]
df = df[df["paw"].str.contains("rf") == False]

dfLH = dfLH[dfLH["paw"].str.contains("rh") == False]
dfLH = dfLH[dfLH["paw"].str.contains("lf") == False]
dfLH = dfLH[dfLH["paw"].str.contains("rf") == False]

dfRH = dfRH[dfRH["paw"].str.contains("lh") == False]
dfRH = dfRH[dfRH["paw"].str.contains("lf") == False]
dfRH = dfRH[dfRH["paw"].str.contains("rf") == False]

dfRF = dfRF[dfRF["paw"].str.contains("lh") == False]
dfRF = dfRF[dfRF["paw"].str.contains("lf") == False]
dfRF = dfRF[dfRF["paw"].str.contains("rh") == False]

df["diff"] = df["run id"].diff()
dfLH["diff"] = dfLH["run id"].diff()
dfRH["diff"] = dfRH["run id"].diff()
dfRF["diff"] = dfRF["run id"].diff()

    
with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\test.xlsx", engine='xlsxwriter') as writer:
      df.to_excel(writer, sheet_name = 'LF')
      dfLH.to_excel(writer, sheet_name = 'LH')
      dfRF.to_excel(writer, sheet_name = 'RF')
      dfRH.to_excel(writer, sheet_name = 'RH')


df2 = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\test.xlsx", sheet_name='LF', header = 0, index_col=None)
df2 = df2.iloc[: , 1:]
df2LH = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\test.xlsx", sheet_name='LH', header = 0, index_col=None)
df2LH = df2LH.iloc[: , 1:]
df2RH = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\test.xlsx", sheet_name='RH', header = 0, index_col=None)
df2RH = df2RH.iloc[: , 1:]
df2RF = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\test.xlsx", sheet_name='RF', header = 0, index_col=None)
df2RF = df2RF.iloc[: , 1:]

index_list = df2.index[df2["diff"] != 0].tolist()
index_listLH = df2LH.index[df2LH["diff"] != 0].tolist()
index_listRH = df2RH.index[df2RH["diff"] != 0].tolist()
index_listRF = df2RF.index[df2RF["diff"] != 0].tolist()


def subtract_from_columns(df2, index_list, value_column):
  for i in range(len(index_list) - 1):
    start_index = index_list[i]
    end_index = index_list[i+1]
    value = df2.at[start_index, value_column]
    df2.iloc[start_index:end_index, 8] -= value
    df2.iloc[start_index:end_index, 9] -= value
  return df2

value_column = 'begintouches'

result = subtract_from_columns(df2,index_list,value_column) 
for i in range(len(result)-1):
    if result["direction"].iloc[i] == 0:
        result["rung"][i] = (result["rung"][i] - 38)*-1
    swingtime.append(result["begintouches"].iloc[i+1]-result["endtouches"].iloc[i])
swingtime.insert(0,0)
result["swingtime"] = swingtime
mask = result['diff'] != 0
result.loc[mask, 'swingtime'] = pd.np.nan
            
resultLH = subtract_from_columns(df2LH,index_listLH,value_column) 
for i in range(len(resultLH)-1):
    if resultLH["direction"].iloc[i] == 0:
        resultLH["rung"][i] = (resultLH["rung"][i] - 38)*-1
    swingtimeLH.append(resultLH["begintouches"].iloc[i+1]-resultLH["endtouches"].iloc[i])
swingtimeLH.insert(0,0)   
resultLH["swingtime"] = swingtimeLH 
mask = resultLH['diff'] != 0
resultLH.loc[mask, 'swingtime'] = pd.np.nan
       
resultRH = subtract_from_columns(df2RH,index_listRH,value_column) 
for i in range(len(resultRH)-1):
    if resultRH["direction"].iloc[i] == 0:
        resultRH["rung"][i] = (resultRH["rung"][i] - 38)*-1
    swingtimeRH.append(resultRH["begintouches"].iloc[i+1]-resultRH["endtouches"].iloc[i])
swingtimeRH.insert(0,0)   
resultRH["swingtime"] = swingtimeRH
mask = resultRH['diff'] != 0
resultRH.loc[mask, 'swingtime'] = pd.np.nan
  
resultRF = subtract_from_columns(df2RF,index_listRF,value_column) 
for i in range(len(resultRF)-1):
    if resultRF["direction"].iloc[i] == 0:
        resultRF["rung"][i] = (resultRF["rung"][i] - 38)*-1
    swingtimeRF.append(resultRF["begintouches"].iloc[i+1]-resultRF["endtouches"].iloc[i])
swingtimeRF.insert(0,0)    
resultRF["swingtime"] = swingtimeRF
mask = resultRF['diff'] != 0
resultRF.loc[mask, 'swingtime'] = pd.np.nan


merge = [result,resultRF,resultRH,resultLH]
merger=pd.concat(merge,axis=0)
merger = merger.reset_index(drop=True)

indexes = []
indexes2= []
for index, row in merger.iterrows():
  if row['diff'] != 0:
    indexes2.append(index)
    
for index, row in merger.iterrows():
  if row['swingtime'] < 0:
    indexes.append(index)

# removeslist = find_range(indexes, indexes2)

# merger = merger.drop(removeslist)
# merger = merger.reset_index(drop=True)

with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\CTR.xlsx", engine='xlsxwriter') as writer:
      result.to_excel(writer, sheet_name = 'LF')
      resultLH.to_excel(writer, sheet_name = 'LH')
      resultRH.to_excel(writer, sheet_name = 'RH')
      resultRF.to_excel(writer, sheet_name = 'RF')
      merger.to_excel(writer, sheet_name = 'merger')

df3 = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\CTR.xlsx", sheet_name='merger', header = 0, index_col=None)
steptime = []
for i in range(len(df3)-1):
    steptime.append(df3["touch"].iloc[i]+df3["swingtime"].iloc[i+1])
steptime.insert(0,0)    
df3["steptime"]=steptime
df3 = df3.iloc[: , 1:]

df3.reset_index()


with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\CTR.xlsx", engine='xlsxwriter') as writer:
      result.to_excel(writer, sheet_name = 'LF')
      resultLH.to_excel(writer, sheet_name = 'LH')
      resultRH.to_excel(writer, sheet_name = 'RH')
      resultRF.to_excel(writer, sheet_name = 'RF')
      merger.to_excel(writer, sheet_name = 'merger')
      df3.to_excel(writer, sheet_name = 'merger+steptime')

df4 = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\CTR.xlsx", sheet_name= 'merger+steptime')
df2 = df4[['diff', 'swingtime','steptime']]
df4 = df4.drop(['diff', 'swingtime','steptime'], axis=1)

df2 =df2[1:]
df2 = df2.append(pd.Series([np.nan]*len(df2.columns), index=df2.columns), ignore_index=True)
df4 = df4.iloc[: , 1:]
df_concatCTR = pd.concat([df4, df2], axis=1)

df_concatCTR['DC'] = df_concatCTR['touch']/df_concatCTR['steptime']
cv2=[]
for i in range(len(df_concatCTR)-1):
    cv2.append(np.abs(2*(df_concatCTR["steptime"].iloc[i+1]-df_concatCTR["steptime"].iloc[i])/(df_concatCTR["steptime"].iloc[i+1]+df_concatCTR["steptime"].iloc[i])))
cv2.insert(-1,0)    
df_concatCTR["cv2"]=cv2

df = df_concatMUT
dff=df_concatCTR

dfLF = df[df['paw'] == 'lf']
dfRF = df[df['paw'] == 'rf']
dfRH = df[df['paw'] == 'rh']
dfLH = df[df['paw'] == 'lh']


dfLF = dfLF.groupby(["subject id","sessionnr"]).median()
dfLH = dfLH.groupby(["subject id","sessionnr"]).median()
dfRF = dfRF.groupby(["subject id","sessionnr"]).median()
dfRH = dfRH.groupby(["subject id","sessionnr"]).median()

columns_to_remove = ["touch", "rung", "begintouches", "endtouches", "sides", "direction", "run id", "diff", "swingtime"]
dfLF = dfLF.drop(columns=columns_to_remove)
dfRF = dfRF.drop(columns=columns_to_remove)
dfRH = dfRH.drop(columns=columns_to_remove)
dfLH = dfLH.drop(columns=columns_to_remove)


test = [dfLF,dfRF,dfLH,dfRH]
final=pd.concat(test,axis=1)
final.columns =["steptimeLF","cv2LF","DCLF","steptimeRF","cv2RF","DCRF","steptimeLH","cv2LH","DCLH","steptimeRH","cv2RH","DCRH"]


dffLF = dff[dff['paw'] == 'lf']
dffRF = dff[dff['paw'] == 'rf']
dffRH = dff[dff['paw'] == 'rh']
dffLH = dff[dff['paw'] == 'lh']


dffLF = dffLF.groupby(["subject id","sessionnr"]).median()
dffLH = dffLH.groupby(["subject id","sessionnr"]).median()
dffRF = dffRF.groupby(["subject id","sessionnr"]).median()
dffRH = dffRH.groupby(["subject id","sessionnr"]).median()

dffLF = dffLF.drop(columns=columns_to_remove)
dffRF = dffRF.drop(columns=columns_to_remove)
dffRH = dffRH.drop(columns=columns_to_remove)
dffLH = dffLH.drop(columns=columns_to_remove)


test2 = [dffLF,dffRF,dffLH,dffRH]
final2=pd.concat(test2,axis=1)
final2.columns =["steptimeLF","cv2LF","DCLF","steptimeRF","cv2RF","DCRF","steptimeLH","cv2LH","DCLH","steptimeRH","cv2RH","DCRH"]


with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\figures\\Sorted\\median_dutycycle_per_session_ercc1_per_mouse2.xlsx", engine='xlsxwriter') as writer:
    final.to_excel(writer, sheet_name = 'MUT')
    final2.to_excel(writer, sheet_name = 'CTR')
#%% PCA this is for 2 groups for the PCA with 4 groups the other 2 groups have to be added manually
df = df_concatCTR
df = df[df["paw"].str.contains("rh") == False]
df = df[df["paw"].str.contains("lh") == False]
df = df[df["paw"].str.contains("lf") == False]

columns_to_include = ['run id', 'touch','sessionnr']
touch_df = df[columns_to_include]

columns_to_include = ['run id', 'steptime','sessionnr']
steptime_df = df[columns_to_include]
steptime_df = steptime_df.dropna(subset=["steptime"])


columns_to_include = ['run id', 'cv2','sessionnr']
cv2_df = df[columns_to_include]
cv2_df = cv2_df.dropna(subset=["cv2"])

columns_to_include = ['run id', 'DC','sessionnr']
DC_df = df[columns_to_include]
DC_df = DC_df.dropna(subset=["DC"])

touch_df = touch_df.groupby(["run id"]).median()
cv2_df = cv2_df.groupby(["run id"]).median()
DC_df = DC_df.groupby(["run id"]).median()
steptime_df = steptime_df.groupby(["run id"]).median()

touch_df = touch_df.merge(cv2_df, how='inner', on='run id').copy()
DC_df = DC_df.merge(cv2_df, how='inner', on='run id').copy()
steptime_df = steptime_df.merge(DC_df, how='inner', on='run id').copy()

df2 = pd.read_excel('C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\control\\Parameters.xlsx')
df2 = df2.drop_duplicates()
df2 = df2.merge(steptime_df, how='inner', on='run id').copy()
df2 = df2.drop(df2.columns[[36,38,40]],axis=1)


df2 = df2.drop(columns=['swing phase rf','swing phase lf','swing phase lh','odd steps','swing phase rh','direction','support 0 paws','state sequence'])
with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\PCA_prep_ctr.xlsx", engine='xlsxwriter') as writer:
      df2.to_excel(writer)
df = df_concatMUT
df = df[df["paw"].str.contains("rh") == False]
df = df[df["paw"].str.contains("lh") == False]
df = df[df["paw"].str.contains("lf") == False]

columns_to_include = ['run id', 'touch','sessionnr']
touch_df = df[columns_to_include]

columns_to_include = ['run id', 'steptime','sessionnr']
steptime_df = df[columns_to_include]
steptime_df = steptime_df.dropna(subset=["steptime"])


columns_to_include = ['run id', 'cv2','sessionnr']
cv2_df = df[columns_to_include]
cv2_df = cv2_df.dropna(subset=["cv2"])

columns_to_include = ['run id', 'DC','sessionnr']
DC_df = df[columns_to_include]
DC_df = DC_df.dropna(subset=["DC"])

touch_df = touch_df.groupby(["run id"]).median()
cv2_df = cv2_df.groupby(["run id"]).median()
DC_df = DC_df.groupby(["run id"]).median()
steptime_df = steptime_df.groupby(["run id"]).median()

touch_df = touch_df.merge(cv2_df, how='inner', on='run id').copy()
DC_df = DC_df.merge(cv2_df, how='inner', on='run id').copy()
steptime_df = steptime_df.merge(DC_df, how='inner', on='run id').copy()

df2 = pd.read_excel('C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\mutant\\Parameters.xlsx')
df2 = df2.drop_duplicates()
df2 = df2.merge(steptime_df, how='inner', on='run id').copy()
df2 = df2.drop(df2.columns[[36,38,40]],axis=1)


df2 = df2.drop(columns=['swing phase rf','swing phase lf','swing phase lh','odd steps','swing phase rh','direction','support 0 paws','state sequence'])
with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\PCA_prep_mut.xlsx", engine='xlsxwriter') as writer:
      df2.to_excel(writer)

ercc1_mut = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\PCA_prep_mut.xlsx")
ercc1_ctr = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\PCA_prep_ctr.xlsx")

Eko_6 = ercc1_mut[ercc1_mut["session nr"] == 6]
Eko_7 = ercc1_mut[ercc1_mut["session nr"] == 7]
Eko_8 = ercc1_mut[ercc1_mut["session nr"] == 8]
Eko_9 = ercc1_mut[ercc1_mut["session nr"] == 9]
Eko_10 = ercc1_mut[ercc1_mut["session nr"] == 10]
Eko_11 = ercc1_mut[ercc1_mut["session nr"] == 11]
Eko_13 = ercc1_mut[ercc1_mut["session nr"] == 13]
Eko_14 = ercc1_mut[ercc1_mut["session nr"] == 14]
Eko_15 = ercc1_mut[ercc1_mut["session nr"] == 15]
Eko_16 = ercc1_mut[ercc1_mut["session nr"] == 16]
Eko_17 = ercc1_mut[ercc1_mut["session nr"] == 17]
Eko_12 = ercc1_mut[ercc1_mut["session nr"] == 12]
Eko_18 = ercc1_mut[ercc1_mut["session nr"] == 18]
Eko_19 = ercc1_mut[ercc1_mut["session nr"] == 19]

# add column to each group with the label before adding them together making one big dataframe
Eko_6["genotype"] = 'Eko_6'
Eko_7["genotype"] = 'Eko_7'
Eko_8["genotype"] = 'Eko_8'
Eko_9["genotype"] = 'Eko_9'
Eko_10["genotype"] = 'Eko_10'
Eko_11["genotype"] = 'Eko_11'
Eko_13["genotype"] = 'Eko_13'
Eko_14["genotype"] = 'Eko_14'
Eko_15["genotype"] = 'Eko_15'
Eko_16["genotype"] = 'Eko_16'
Eko_17["genotype"] = 'Eko_17'
Eko_12["genotype"] = 'Eko_12'
Eko_18["genotype"] = 'Eko_18'
Eko_19["genotype"] = 'Eko_19'

Ewt_6 = ercc1_ctr[ercc1_ctr["session nr"] == 6]
Ewt_7 = ercc1_ctr[ercc1_ctr["session nr"] == 7]
Ewt_8 = ercc1_ctr[ercc1_ctr["session nr"] == 8]
Ewt_9 = ercc1_ctr[ercc1_ctr["session nr"] == 9]
Ewt_10 = ercc1_ctr[ercc1_ctr["session nr"] == 10]
Ewt_11 = ercc1_ctr[ercc1_ctr["session nr"] == 11]
Ewt_13 = ercc1_ctr[ercc1_ctr["session nr"] == 13]
Ewt_14 = ercc1_ctr[ercc1_ctr["session nr"] == 14]
Ewt_15 = ercc1_ctr[ercc1_ctr["session nr"] == 15]
Ewt_16 = ercc1_ctr[ercc1_ctr["session nr"] == 16]
Ewt_17 = ercc1_ctr[ercc1_ctr["session nr"] == 17]
Ewt_12 = ercc1_ctr[ercc1_ctr["session nr"] == 12]
Ewt_18 = ercc1_ctr[ercc1_ctr["session nr"] == 18]
Ewt_19 = ercc1_ctr[ercc1_ctr["session nr"] == 19]

# add column to each group with the label before adding them together making one big dataframe
Ewt_6["genotype"] = 'Ewt_6'
Ewt_7["genotype"] = 'Ewt_7'
Ewt_8["genotype"] = 'Ewt_8'
Ewt_9["genotype"] = 'Ewt_9'
Ewt_10["genotype"] = 'Ewt_10'
Ewt_11["genotype"] = 'Ewt_11'
Ewt_13["genotype"] = 'Ewt_13'
Ewt_14["genotype"] = 'Ewt_14'
Ewt_15["genotype"] = 'Ewt_15'
Ewt_16["genotype"] = 'Ewt_16'
Ewt_17["genotype"] = 'Ewt_17'
Ewt_12["genotype"] = 'Ewt_12'
Ewt_18["genotype"] = 'Ewt_18'
Ewt_19["genotype"] = 'Ewt_19'

# add column to each group with the label before adding them together making one big dataframe


groups = [Ewt_6, Ewt_7, Ewt_8,Ewt_9,Ewt_10,Ewt_11,Ewt_12,Ewt_13,Ewt_14,Ewt_15,Ewt_16,Ewt_17,Ewt_18,Ewt_19,Eko_6, Eko_7, Eko_8,Eko_9,Eko_10,Eko_11,Eko_12,Eko_13,Eko_14,Eko_15,Eko_16,Eko_17,Eko_18,Eko_19]

all_groups = pd.concat(groups, ignore_index= True)

features = all_groups[all_groups.columns[5:31]]

labels = all_groups["genotype"]
sessions = all_groups["session nr"]
subjects = all_groups["subject id"]

features_norm = StandardScaler().fit_transform(features)
features_norm = pd.DataFrame(features_norm, columns= features.columns)

pca = PCA(n_components = 0.85)


principal_comp = pca.fit_transform(features_norm)
principal_df = pd.DataFrame(data = principal_comp)
final_df = pd.concat([principal_df, labels,sessions,subjects], axis = 1)

final_df=final_df.groupby(["genotype"]).median().reset_index()
with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\scatterplotPCAses6-12-18SCA1+ercc1.xlsx", engine='xlsxwriter') as writer:
    final_df.to_excel(writer)
pca.explained_variance_ratio_
# scree plot
pc_values = np.arange(pca.n_components_) + 1
plt.plot(pc_values, pca.explained_variance_ratio_, 'ro-')
plt.savefig("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\PCA\\EVratio.eps")
plt.show()

print ("Proportion of Variance Explained : ", pca.explained_variance_ratio_)    
out_sum = np.cumsum(pca.explained_variance_ratio_)  
print ("Cumulative Prop. Variance Explained: ", out_sum)
# plot pca
fig = plt.figure(figsize = (8,8))
ax = fig.add_subplot(1,1,1)
ax.set_xlabel('Principal Component 1', fontsize = 15)
ax.set_ylabel('Principal Component 2', fontsize = 15)
ax.set_title('2 component PCA', fontsize = 20)

targets = ['Eko_6', 'Eko_7','Eko_8','Eko_9','Eko_10','Eko_11','Eko_12','Eko_13','Eko_14','Eko_15','Eko_16','Eko_17', 'Eko_18','Eko_19','Ewt_6', 'Ewt_7','Ewt_8','Ewt_9','Ewt_10','Ewt_11','Ewt_12','Ewt_13','Ewt_14','Ewt_15','Ewt_16','Ewt_17', 'Ewt_18','Ewt_19']


colors = ['#F98B88','#F75D59','#FF6347','#FF0000','#F70D1A','#F62817','#DC381F','#C11B17','#9F000F','#8B0000','#8C001A','#7E191B','#800517','#660000','#16E2F5','#0AFFFF','#57FEFF','#9AFEFF','#AFDCEC','#82CAFF','#00BFFF','#6495ED','#1E90FF','#1E90FF','#2B60DE','#2554C7','#0909FF','#0000A5']


for target, color in zip(targets, colors):
    indicesToKeep = final_df["genotype"] == target
    points = ax.scatter(final_df.loc[indicesToKeep, 0], final_df.loc[indicesToKeep, 1], c = color, s = 50)
# ax.legend(targets, loc='lower right',fontsize=6)
ax.grid()
# ax.set(xlim=(-2, 3.5), ylim=(-1.5, 2))
plt.savefig("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\PCA\\finalscatterplot.eps", format="eps")
components = pca.components_

# Create a DataFrame from the components array
df = pd.DataFrame(components, columns=['time run','N.o.Steps','Even steps','step2','step4','step6','stance phase lf','stance phase lh','stance phase rf','stance phase rh','lag lf/rh','lag rf/lh','support diagonal','support girlde','support lateral','support 1 paw','support 3 paws','support 4paws','high rungs front','high rungs hind','low rungs front','low rungs hind','mean touchtime','steptime','DC','cv2'])

# Write the DataFrame to an Excel file
df.to_excel("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\PCA\\FeatureperCompinent.xlsx")

# features_train, features_test, labels_train, labels_test = train_test_split(features_norm, labels, test_size = 0.25, stratify = labels)

# # create model
# forest = RandomForestClassifier(n_estimators = 20000, max_features = 'auto', max_samples= None)

# # train model
# forest.fit(features_train, labels_train)

# # prediction
# pred = forest.predict(features_test)
# print("Accuracy:", metrics.accuracy_score(labels_test, pred))
with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\PCA\\PCA.xlsx", engine='xlsxwriter') as writer:
    final_df.to_excel(writer)
#%% heatmaps The heatmaps that are made with the code are OF 1 PAIR OF PAWS 
#if you want all 3 pairs run this part of the code multiple times 
#and alter the variables to get different combinations of paws!!!!
swingtime=[]
# dfdiff=[]
df = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\control\\alltouches.xlsx")
df = df.iloc[: , 1:]

df = df[df["paw"].str.contains("rh") == False]
df = df[df["paw"].str.contains("lf") == False]
df2=df
for i in range(len(df)-1):
    if (df['paw'].iloc[i] == df['paw'].iloc[i+1]):
        df2 = df2.drop(df.index[i])
    else:
        i=i+1

df2["diff"] = df2["run id"].diff()


    
with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\figures\\sorted\\ctr_testje.xlsx", engine='xlsxwriter') as writer:
      df2.to_excel(writer)



df2 = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\figures\\sorted\\ctr_testje.xlsx", header = 0, index_col=None)
df2 = df2.iloc[: , 1:]


index_list = df2.index[df2["diff"] != 0].tolist()



def subtract_from_columns(df2, index_list, value_column):
  for i in range(len(index_list) - 1):
    start_index = index_list[i]
    end_index = index_list[i+1]
    value = df2.at[start_index, value_column]
    df2.iloc[start_index:end_index, 8] -= value
    df2.iloc[start_index:end_index, 9] -= value
  return df2

value_column = 'begintouches'

result = subtract_from_columns(df2,index_list,value_column) 
for i in range(len(result)-2):
    if result["direction"].iloc[i] == 0:
        result["rung"][i] = (result["rung"][i] - 38)*-1
    swingtime.append(result["begintouches"].iloc[i+2]-result["endtouches"].iloc[i])
swingtime.insert(0,0)
swingtime.append(0)
result["swingtime"] = swingtime
mask = result['diff'] != 0
mask2 = result['swingtime']<=0
result.loc[mask, 'swingtime'] = pd.np.nan
result.loc[mask2, 'swingtime'] = pd.np.nan


merge = [result]
merger=pd.concat(merge,axis=0)
merger = merger.reset_index(drop=True)

indexes = []
indexes2= []
for index, row in merger.iterrows():
  if row['diff'] != 0:
    indexes2.append(index)
    
for index, row in merger.iterrows():
  if row['swingtime'] < 0:
    indexes.append(index)

# removeslist = find_range(indexes, indexes2)

# merger = merger.drop(removeslist)
# merger = merger.reset_index(drop=True)

with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\figures\\Sorted\\CTR_RFLH.xlsx", engine='xlsxwriter') as writer:
      merger.to_excel(writer)



df3 = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\figures\\Sorted\\CTR_RFLH.xlsx", header = 0, index_col=None)
steptime = []

for i in range(len(df3)-1):
    steptime.append(df3["touch"].iloc[i]+df3["swingtime"].iloc[i+1])
steptime.insert(0,0) 
df3["steptime"]=steptime
df3 = df3.iloc[: , 1:]

df3.reset_index()
delay = []
duty_cycle=[]
for i in range(len(df3)-1):
    delay.append((df3["begintouches"].iloc[i+1]-df3["begintouches"].iloc[i])/df3["steptime"].iloc[i+1])
    duty_cycle.append((df3["touch"].iloc[i])/df3["steptime"].iloc[i+1])
delay.insert(0,0) 
duty_cycle.insert(0,0) 
df3["delay"]=delay
df3["duty cycle"]=duty_cycle
df2 = df3[['delay', 'swingtime','steptime','duty cycle']]
df3 = df3.drop(['duty cycle','delay', 'swingtime','steptime'], axis=1)

df2 =df2[1:]
df2 = df2.append(pd.Series([np.nan]*len(df2.columns), index=df2.columns), ignore_index=True)
df3 = df3.iloc[: , 1:]
df_CTR = pd.concat([df3, df2], axis=1)

with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\CTR_RFLH.xlsx", engine='xlsxwriter') as writer:
      df_CTR.to_excel(writer)
      
swingtime=[]
# dfdiff=[]
df = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\mutant\\alltouches.xlsx")
df = df.iloc[: , 1:]

df = df[df["paw"].str.contains("rh") == False]
df = df[df["paw"].str.contains("lf") == False]
df2=df
for i in range(len(df)-1):
    if (df['paw'].iloc[i] == df['paw'].iloc[i+1]):
        df2 = df2.drop(df.index[i])
    else:
        i=i+1

df2["diff"] = df2["run id"].diff()


    
with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\figures\\Sorted\\mut_testje.xlsx", engine='xlsxwriter') as writer:
      df2.to_excel(writer)



df2 = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\figures\\Sorted\\mut_testje.xlsx", header = 0, index_col=None)
df2 = df2.iloc[: , 1:]


index_list = df2.index[df2["diff"] != 0].tolist()



def subtract_from_columns(df2, index_list, value_column):
  for i in range(len(index_list) - 1):
    start_index = index_list[i]
    end_index = index_list[i+1]
    value = df2.at[start_index, value_column]
    df2.iloc[start_index:end_index, 8] -= value
    df2.iloc[start_index:end_index, 9] -= value
  return df2

value_column = 'begintouches'

result = subtract_from_columns(df2,index_list,value_column) 
for i in range(len(result)-2):
    if result["direction"].iloc[i] == 0:
        result["rung"][i] = (result["rung"][i] - 38)*-1
    swingtime.append(result["begintouches"].iloc[i+2]-result["endtouches"].iloc[i])
swingtime.insert(0,0)
swingtime.append(0)
result["swingtime"] = swingtime
mask = result['diff'] != 0
mask2 = result['swingtime']<=0
result.loc[mask, 'swingtime'] = pd.np.nan
result.loc[mask2, 'swingtime'] = pd.np.nan


merge = [result]
merger=pd.concat(merge,axis=0)
merger = merger.reset_index(drop=True)

indexes = []
indexes2= []
for index, row in merger.iterrows():
  if row['diff'] != 0:
    indexes2.append(index)
    
for index, row in merger.iterrows():
  if row['swingtime'] < 0:
    indexes.append(index)


with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\figures\\Sorted\\MUT_RFLH.xlsx", engine='xlsxwriter') as writer:
      merger.to_excel(writer)



df3 = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\figures\\Sorted\\MUT_RFLH.xlsx")
steptime = []

for i in range(len(df3)-1):
    steptime.append(df3["touch"].iloc[i]+df3["swingtime"].iloc[i+1])
steptime.insert(0,0) 
df3["steptime"]=steptime
df3 = df3.iloc[: , 1:]

df3.reset_index()
delay = []
duty_cycle=[]
for i in range(len(df3)-1):
    delay.append((df3["begintouches"].iloc[i+1]-df3["begintouches"].iloc[i])/df3["steptime"].iloc[i+1])
    duty_cycle.append((df3["touch"].iloc[i])/df3["steptime"].iloc[i+1])
delay.insert(0,0) 
duty_cycle.insert(0,0) 
df3["delay"]=delay
df3["duty cycle"]=duty_cycle
df2 = df3[['delay', 'swingtime','steptime','duty cycle']]
df3 = df3.drop(['duty cycle','delay', 'swingtime','steptime'], axis=1)

df2 =df2[1:]
df2 = df2.append(pd.Series([np.nan]*len(df2.columns), index=df2.columns), ignore_index=True)
df3 = df3.iloc[: , 1:]
df_MUT = pd.concat([df3, df2], axis=1)

with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\MUT_RFLH.xlsx", engine='xlsxwriter') as writer:
      df_MUT.to_excel(writer)



df = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\figures\\Sorted\\CTR_RFLH.xlsx", header=0, index_col=None)
df = df[(df.delay >= 0) & (df.delay <= 1) & df["paw"].eq("rf")].dropna(subset=["delay"])
df = df.iloc[:,1:]

dfs = []
for i in range(6, 20):
    dfs.append(df[df["sessionnr"] == i])

df2 = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\figures\\Sorted\\MUT_RFLH.xlsx", header=0, index_col=None)
df2 = df2[(df2.delay >= 0) & (df2.delay <= 1) & df2["paw"].eq("rf")].dropna(subset=["delay"])
df2 = df2.iloc[:,1:]
for i in range(6, 20):
    dfs.append(df2[df2["sessionnr"] == i])

# Create a Pandas Excel writer using XlsxWriter as the engine
writer = pd.ExcelWriter('C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\heatmaps\\HBall_RFLH.xlsx', engine='xlsxwriter')

# Iterate over the dataframes and write each to a separate sheet
for df in dfs:
    sheet_name = f"{df['subject id'].iloc[0]} - Session {df['sessionnr'].iloc[0]}"
    df.to_excel(writer, sheet_name=sheet_name, index=False)

# Save the Excel file
writer.save()
#%% old version of the code from line 1174 to 1197   
# df = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\figures\\test\\all analysis\\CTR_RFLH.xlsx", header = 0, index_col=None)
# df = df.drop(df[df.delay < 0].index)
# df = df.drop(df[df.delay > 1].index)

# df = df.dropna(subset=["delay"])

# df6= df[df["sessionnr"]==6]
# df6rf = df6[df6["paw"]=='rf']

# df7= df[df["sessionnr"]==7]
# df7rf = df7[df7["paw"]=='rf']

# df8= df[df["sessionnr"]==8]
# df8rf = df8[df8["paw"]=='rf']

# df9= df[df["sessionnr"]==9]
# df9rf = df9[df9["paw"]=='rf']

# df10= df[df["sessionnr"]==10]
# df10rf = df10[df10["paw"]=='rf']

# df11= df[df["sessionnr"]==11]
# df11rf = df11[df11["paw"]=='rf']

# df12= df[df["sessionnr"]==12]
# df12rf = df12[df12["paw"]=='rf']

# df13= df[df["sessionnr"]==13]
# df13rf = df13[df13["paw"]=='rf']

# df14= df[df["sessionnr"]==14]
# df14rf = df14[df14["paw"]=='rf']

# df15= df[df["sessionnr"]==15]
# df15rf = df15[df15["paw"]=='rf']

# df16= df[df["sessionnr"]==16]
# df16rf = df16[df16["paw"]=='rf']

# df17= df[df["sessionnr"]==17]
# df17rf = df17[df17["paw"]=='rf']


# df18 = df[df["sessionnr"]==18]
# df18rf = df18[df18["paw"]=='rf']

# df19 = df[df["sessionnr"]==19]
# df19rf = df19[df19["paw"]=='rf']

# dfM = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\Figures\\test\\all analysis\\MUT_RFLH.xlsx", header = 0, index_col=None)
# dfM = dfM.drop(dfM[dfM.delay < 0].index)
# dfM = dfM.drop(dfM[dfM.delay > 1].index)

# dfM=dfM.dropna(subset=['delay'])

# dfM6= dfM[dfM["sessionnr"]==6]
# dfM6rf = dfM6[dfM6["paw"]=='rf']

# dfM7= dfM[dfM["sessionnr"]==7]
# dfM7rf = dfM7[dfM7["paw"]=='rf']

# dfM8= dfM[dfM["sessionnr"]==8]
# dfM8rf = dfM8[dfM8["paw"]=='rf']

# dfM9= dfM[dfM["sessionnr"]==9]
# dfM9rf = dfM9[dfM9["paw"]=='rf']

# dfM10= dfM[dfM["sessionnr"]==10]
# dfM10rf = dfM10[dfM10["paw"]=='rf']

# dfM11= dfM[dfM["sessionnr"]==11]
# dfM11rf = dfM11[dfM11["paw"]=='rf']

# dfM12= dfM[dfM["sessionnr"]==12]
# dfM12rf = dfM12[dfM12["paw"]=='rf']

# dfM13= dfM[dfM["sessionnr"]==13]
# dfM13rf = dfM13[dfM13["paw"]=='rf']

# dfM14= dfM[dfM["sessionnr"]==14]
# dfM14rf = dfM14[dfM14["paw"]=='rf']

# dfM15= dfM[dfM["sessionnr"]==15]
# dfM15rf = dfM15[dfM15["paw"]=='rf']

# dfM16= dfM[dfM["sessionnr"]==16]
# dfM16rf = dfM16[dfM16["paw"]=='rf']

# dfM17= dfM[dfM["sessionnr"]==17]
# dfM17rf = dfM17[dfM17["paw"]=='rf']


# dfM18 = dfM[dfM["sessionnr"]==18]
# dfM18rf = dfM18[dfM18["paw"]=='rf']

# dfM19 = dfM[dfM["sessionnr"]==19]
# dfM19rf = dfM19[dfM19["paw"]=='rf']

# with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\Figures\\test\\all analysis\\heatmaps\\HBall_RFLH.xlsx", engine='xlsxwriter') as writer:
#       df6rf.to_excel(writer, sheet_name = 'CTR6rf')
#       df7rf.to_excel(writer, sheet_name = 'CTR7rf')
#       df8rf.to_excel(writer, sheet_name = 'CTR8rf')
#       df9rf.to_excel(writer, sheet_name = 'CTR9rf')
#       df10rf.to_excel(writer, sheet_name = 'CTR10rf')
#       df11rf.to_excel(writer, sheet_name = 'CTR11rf')
#       df12rf.to_excel(writer, sheet_name = 'CTR12rf')
#       df13rf.to_excel(writer, sheet_name = 'CTR13rf')
#       df14rf.to_excel(writer, sheet_name = 'CTR14rf')
#       df15rf.to_excel(writer, sheet_name = 'CTR15rf')
#       df16rf.to_excel(writer, sheet_name = 'CTR16rf')
#       df17rf.to_excel(writer, sheet_name = 'CTR17rf')
#       df18rf.to_excel(writer, sheet_name = 'CTR18rf')
#       df19rf.to_excel(writer, sheet_name = 'CTR19rf')
#       dfM6rf.to_excel(writer, sheet_name = 'MUT6rf')
#       dfM7rf.to_excel(writer, sheet_name = 'MUT7rf')
#       dfM8rf.to_excel(writer, sheet_name = 'MUT8rf')
#       dfM9rf.to_excel(writer, sheet_name = 'MUT9rf')
#       dfM10rf.to_excel(writer, sheet_name = 'MUT10rf')
#       dfM11rf.to_excel(writer, sheet_name = 'MUT11rf')
#       dfM12rf.to_excel(writer, sheet_name = 'MUT12rf')
#       dfM13rf.to_excel(writer, sheet_name = 'MUT13rf')
#       dfM14rf.to_excel(writer, sheet_name = 'MUT14rf')
#       dfM15rf.to_excel(writer, sheet_name = 'MUT15rf')
#       dfM16rf.to_excel(writer, sheet_name = 'MUT16rf')      
#       dfM17rf.to_excel(writer, sheet_name = 'MUT17rf')
#       dfM18rf.to_excel(writer, sheet_name = 'MUT18rf')
#       dfM19rf.to_excel(writer, sheet_name = 'MUT19rf')
#%%
import openpyxl
allvars = []
allsheetnames=[]
file = "C:\\LabProject\\LadderProject\\Data\\figures\\Sorted\\heatmaps\\HBall_RFLH.xlsx"
workbook = openpyxl.load_workbook(file)
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows():
        data.append([cell.value for cell in row])
    df = pd.DataFrame(data)
    x = df[13]
    y = df[10]
    xlist = x.tolist()
    ylist = y.tolist()
    xlist = xlist[2:]
    ylist = ylist[2:]
    array=np.array([xlist,ylist])
    std_dev = np.round(np.std(array),4)
    allsheetnames.append(sheet_name)
    print(sheet_name)
    print(std_dev)
    scaling = len(xlist)*0.067
    xlist = [x * 100 for x in xlist]
    ylist = [y * 100 for y in ylist]
    plt.hist2d(xlist, ylist, bins = ([np.arange(0, 105, 5), np.arange(0, 105, 5)]), vmin=0, vmax=scaling, cmap = 'Greys')
    # plt.axis('off')
    plt.title(sheet_name)
    plt.colorbar()
    plt.savefig('C:\\LabProject\\LadderProject\\Data\\figures\\Sorted\\heatmaps\\'+sheet_name+'.Tiff', format='Tiff',bbox_inches='tight')
    plt.show()

#%% circularity/ excentricity 
df = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\mutant\\Lag.xlsx")
df2 = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\control\\Lag.xlsx")
df = df[df["pair"]=="rf/lh"]
df2 = df2[df2["pair"]=="rf/lh"]

df["lag"] = df["lag"].div(2)
df2["lag"]=df2["lag"].div(2)
df["lag"] = round(df["lag"],1)
df2["lag"] = round(df2["lag"],1)

data_frames = {}
data_frames2={}
for i in range(6, 20):
    data_frames[f"ko_{i}"] = df[df["ses nr"] == i]
    data_frames2[f"wt_{i}"] = df2[df2["ses nr"] == i]

    
    data_frames[f"ko_{i}"] = data_frames[f"ko_{i}"].groupby(["subject id"])["lag"].value_counts()
    data_frames2[f"wt_{i}"] = data_frames2[f"wt_{i}"].groupby(["subject id"])["lag"].value_counts()
dfmut = pd.concat(data_frames,axis=1)
dfctr = pd.concat(data_frames2,axis=1)


with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\lag(10bins).xlsx", engine='xlsxwriter') as writer:
      dfmut.to_excel(writer, sheet_name = 'KO')
      dfctr.to_excel(writer, sheet_name = 'WT')


import shapely.geometry as geometry
def pol2cart(theta, rho):
    x = rho * np.cos(theta)
    y = rho * np.sin(theta)
    return x, y


from shapely.geometry import Point


# Define the input and output file paths
input_file = "C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\lag(10bins).xlsx"
output_file = "C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\circularity_and_centroidsWT.xlsx"

# Define a list of subject IDs to process
subject_ids = [751,752,759,762,753,754,755,809,807,804,802]

# Create an Excel writer object to write the results to the output file
writer = pd.ExcelWriter(output_file)

for subject_id in subject_ids:

    # Load the data for the current subject ID
    df = pd.read_excel(input_file, sheet_name='WT')
    
    for i in range(len(df)):
        # check if the subject id is NaN
        if pd.isna(df.loc[i, "subject id"]):
            # replace NaN with the subject id from the row above
            df.loc[i, "subject id"] = df.loc[i-1, "subject id"]
    df= df.fillna(0)

    df_subject = df[df["subject id"] == subject_id]
    rows_to_sum = df_subject.loc[df_subject['lag'].isin([0, 1])]
     
    row_sum = rows_to_sum.sum()
    df_subject.loc[df_subject.index.min()] = row_sum
    df_subject = df_subject.drop(rows_to_sum.index)
    df_subject.loc[df_subject['lag'] == 1, 'lag'] = 0
    df_subject = df_subject.sort_values(by=["lag"])

    circularlist = []
    centroidlist = []

    for i in range(6,20):

         data = df_subject[f'wt_{i}'].tolist()
         data.append(data[0])
         label_loc = np.linspace(start=0, stop=2 * np.pi, num=len(data))
         x2, y2 = pol2cart(label_loc, data)
         polygon = geometry.Polygon(zip(x2, y2))
         area = polygon.area
         x, y = pol2cart(label_loc, data)
         polygon = geometry.Polygon(zip(x, y))
         perimeter = polygon.length
         if perimeter ==0:
             continue
         circularity = 4 * np.pi * area / perimeter ** 2
         circularity = round(circularity, 3)
         circularlist.append(circularity)
         COG = polygon.centroid
         centroid_distance = Point(0, 0).distance(COG)
         centroidlist.append(centroid_distance)

         plt.figure(figsize=(8, 8))
         plt.subplot(polar=True)
         plt.plot(label_loc, data, label='data')
         plt.legend()
         plt.show()

     # Create a DataFrame to store the results
    results_df = pd.DataFrame({"Circularity": circularlist, "Centroid Distance": centroidlist})

     # Write the results to a sheet with the current subject ID
    results_df.to_excel(writer, sheet_name=f"Subject {subject_id}", index=False)

 # Save and close the Excel file
writer.save()

input_file = "C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\lag(10bins).xlsx"
output_file = "C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\circularity_and_centroidsKO.xlsx"

# Define a list of subject IDs to process
subject_ids = [756,757,758,760,761,801,803,805,806,808,810]

# Create an Excel writer object to write the results to the output file
writer = pd.ExcelWriter(output_file)

for subject_id in subject_ids:

    # Load the data for the current subject ID
    df = pd.read_excel(input_file, sheet_name='KO')
    for i in range(len(df)):
        # check if the subject id is NaN
        if pd.isna(df.loc[i, "subject id"]):
            # replace NaN with the subject id from the row above
            df.loc[i, "subject id"] = df.loc[i-1, "subject id"]
    df= df.fillna(0)
    df_subject = df[df["subject id"] == subject_id]
    rows_to_sum = df_subject.loc[df_subject['lag'].isin([0, 1])]
    
    row_sum = rows_to_sum.sum()
    df_subject.loc[df_subject.index.min()] = row_sum

    df_subject = df_subject.drop(rows_to_sum.index)
    df_subject.loc[df_subject['lag'] == 1, 'lag'] = 0
    df_subject = df_subject.sort_values(by=["lag"])

    circularlist = []
    centroidlist = []

    for i in range(6,20):

        data = df_subject[f'ko_{i}'].tolist()
        data.append(data[0])
        label_loc = np.linspace(start=0, stop=2 * np.pi, num=len(data))
        x2, y2 = pol2cart(label_loc, data)
        polygon = geometry.Polygon(zip(x2, y2))
        area = polygon.area
        x, y = pol2cart(label_loc, data)
        polygon = geometry.Polygon(zip(x, y))
        perimeter = polygon.length
        if perimeter ==0:
            continue
        circularity = 4 * np.pi * area / perimeter ** 2
        circularity = round(circularity, 3)
        circularlist.append(circularity)
        COG = polygon.centroid
        centroid_distance = Point(0, 0).distance(COG)
        centroidlist.append(centroid_distance)

        plt.figure(figsize=(8, 8))
        plt.subplot(polar=True)
        plt.plot(label_loc, data, label='data')
        plt.legend()
        plt.show()

    # Create a DataFrame to store the results
    results_df = pd.DataFrame({"Circularity": circularlist, "Centroid Distance": centroidlist})

    # Write the results to a sheet with the current subject ID
    results_df.to_excel(writer, sheet_name=f"Subject {subject_id}", index=False)

# Save and close the Excel file
writer.save()
#%% support circles 
ercc1_mut = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\mutant\\Parameters.xlsx")
ercc1_ctr = pd.read_excel("C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\control\\Parameters.xlsx")
df = pd.DataFrame(ercc1_ctr)
df2 = pd.DataFrame(ercc1_mut)



ctr = df.groupby(["session nr","subject id"]).median()
ctr = ctr.drop(["direction", "timerun", "number of steps", "even steps", "odd steps","run id", "step 6", "step 2", "step 4", "swing phase lh", "swing phase rh","swing phase lf", "stance phase rf", "stance phase lf", "stance phase rh","stance phase lh", "lag lf/rh", "lag rf/lh", "swing phase rf", "session id","high rungs front", "high rungs hind", "low rungs front", "low rungs hind"], axis=1)



mut = df2.groupby(["session nr","subject id"]).median()
mut = mut.drop(["direction", "timerun", "number of steps", "even steps", "odd steps","run id", "step 6", "step 2", "step 4", "swing phase lh", "swing phase rh",                "swing phase lf", "stance phase rf", "stance phase lf", "stance phase rh",                "stance phase lh", "lag lf/rh", "lag rf/lh", "swing phase rf", "session id",                "high rungs front", "high rungs hind", "low rungs front", "low rungs hind"], axis=1)



with pd.ExcelWriter("C:\\LabProject\\LadderProject\\Data\\Figures\\Sorted\\supports.xlsx", engine='xlsxwriter') as writer:
    mut.to_excel(writer, sheet_name='mut')
    ctr.to_excel(writer, sheet_name='ctr')
#%%
