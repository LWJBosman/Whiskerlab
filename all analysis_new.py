import pandas as pd 
import seaborn as sns
import matplotlib as plt
import matplotlib.pyplot as plt 
import numpy as np
from sklearn.decomposition import PCA
from shapely.geometry import Point

from sklearn.preprocessing import StandardScaler

#%% manually enetered values
#input files base path
base_path = 'C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\'

#output files base path
base_path_out = 'C:\\LabProject\\LadderProject\\Data\\excel\\Sorted\\results\\'
#in the output path create 3 folders named RFLH RFLF RFRH on your device these will be used later.

#first and last session number you want to analyze 
first_session_number = 6
last_session_number = 19

#subject ids of the mice you want to analyze
subject_ids_KO = [190,191,193,196,198,201,203,205]
subject_ids_WT = [189,194,195,197,199,200,202,204]
#%% overview of all the files the script uses/creates
#input files
lag_file_mut = base_path + 'mutant\\Lag.xlsx'
lag_file_ctr = base_path + 'control\\Lag.xlsx'
parameter_file_mut = base_path + 'mutant\\Parameters.xlsx'
parameter_file_ctr = base_path + 'control\\Parameters.xlsx'
alltouches_file_mut = base_path + 'mutant\\alltouches.xlsx'
alltouches_file_ctr = base_path + 'control\\alltouches.xlsx'
steps_front_file_mut = base_path + 'mutant\\steps_front.xlsx'
steps_front_file_ctr = base_path + 'control\\steps_front.xlsx'

#output files
median_runtime_per_session_per_mouse = base_path_out + 'median_runtime_per_session_per_mouse.xlsx' 
median_touchtime_per_session_per_mouse = base_path_out + 'median_touchtime_per_session_per_mouse.xlsx'
step_type_per_mouse_per_session = base_path_out + 'step_type_per_mouse_per_session.xlsx'
runtypes_mut = base_path_out + 'runtypes_mut.xlsx'
runtypes_ctr= base_path_out + 'runtypes_ctr.xlsx'
median_DC_CV2_Steptime = base_path_out + 'median_DC+CV2+steptime_per_paw'
scatterplot_pca = base_path_out + 'scatterplotPCAses6-12-18SCA1+ercc1.xlsx'
feature_per_component_pca = base_path_out + 'FeatureperCompinent.xlsx'
PCA_values = base_path_out + 'PCA.xlsx'
HBall_RFLH = base_path_out + 'HBall_RFLH.xlsx'
HBall_RFLF = base_path_out + 'HBall_RFLF.xlsx'
HBall_RFRH = base_path_out + 'HBall_RFRH.xlsx'

lag_10bins = base_path_out + 'lag(10bins).xlsx'
circularity_and_centroidsWT = base_path_out + 'circularity_and_centroidsWT.xlsx'
circularity_and_centroidsKO = 'circularity_and_centroidsKO.xlsx'

#output figures
runtime = base_path_out + 'runtimes.eps'
touchtime_alltouches = base_path_out + 'touchtime all touches.eps'
EV_ratio=base_path_out + 'EVratio.eps'
PCA_plot =base_path_out + 'finalscatterplot.eps'
supports = base_path_out + 'supports.xlsx'
heatmapsRFLH = base_path_out + 'RFLH'
heatmapsRFRH = base_path_out + 'RFRH'
heatmapsRFLF = base_path_out + 'RFLF'

#inbetween files used as checkpoints to store some of the data
test_file = base_path_out + 'test.xlsx'
MUT_inbetween_file = base_path_out + 'MUT.xlsx'
CTR_inbetween_file = base_path_out + 'CTR.xlsx'
Pca_prep_ctr = base_path_out + 'PCA_prep_ctr.xlsx'
Pca_prep_mut = base_path_out + 'PCA_prep_mut.xlsx'

heatmaps_ctr_inbetween = base_path_out + 'ctr_testje.xlsx'
heatmaps_mut_inbetween = base_path_out + 'mut_testje.xlsx'
heatmaps_ctr_RFLH = base_path_out + 'CTR_RFLH.xlsx'
heatmaps_mut_RFLH = base_path_out + 'MUT_RFLH.xlsx'
heatmaps_ctr_RFLF = base_path_out + 'CTR_RFLF.xlsx'
heatmaps_mut_RFLF = base_path_out + 'MUT_RFLF.xlsx'
heatmaps_ctr_RFRH = base_path_out + 'CTR_RFRH.xlsx'
heatmaps_mut_RFRH = base_path_out + 'MUT_RFRH.xlsx'
amount_of_sessions= last_session_number - first_session_number +1
middle_session_number = round(((last_session_number-first_session_number)/2)+first_session_number)
#%% start of the analysis code 
#median trial times per mouse per session
# this part of the code calculates the median trialtime per mouse per session from the parameter 
# excel file and writes the results into a new excel file
ercc1_mut = pd.read_excel(parameter_file_mut)
ercc1_ctr = pd.read_excel(parameter_file_ctr)


df2 = pd.DataFrame(ercc1_mut)
dff2 = df2.groupby(["subject id","session nr"]).median()
dff2= dff2["timerun"]
df = pd.DataFrame(ercc1_ctr)
dff = df.groupby(["subject id","session nr"]).median()
dff= dff["timerun"]

test = [dff,dff2]
result = pd.concat(test, axis=1)
with pd.ExcelWriter(median_runtime_per_session_per_mouse, engine='xlsxwriter') as writer:
    result.to_excel(writer)
    
#%%median touchtime per mouser per session
# this part of the code calculates the median touchtime per mouse per session from the alltouches 
# excel file and writes the results into a new excel file
ercc1_mut = pd.read_excel(alltouches_file_mut)
ercc1_ctr = pd.read_excel(alltouches_file_ctr)

# filter for high rung touches
if (ercc1_mut["sides"] == 1).all():
    ercc1_mut=ercc1_mut.drop(ercc1_mut[ercc1_mut.rung%2!=0].index)
if (ercc1_mut["sides"] == 0).all():
    ercc1_mut=ercc1_mut.drop(ercc1_mut[ercc1_mut.rung%2==0].index)
if (ercc1_ctr["sides"] == 1).all():
    ercc1_ctr=ercc1_ctr.drop(ercc1_ctr[ercc1_ctr.rung%2!=0].index)
if (ercc1_ctr["sides"] == 0).all():
    ercc1_ctr=ercc1_ctr.drop(ercc1_ctr[ercc1_ctr.rung%2==0].index)

df2 = pd.DataFrame(ercc1_mut)
df2.pop("rung")
dff2 = df2.groupby(["subject id","sessionnr"]).median()
dff2= dff2["touch"]
df = pd.DataFrame(ercc1_ctr)
df.pop("rung")
dff = df.groupby(["subject id","sessionnr"]).median()
dff= dff["touch"]

test = [dff, dff2]
result = pd.concat(test, axis=1)
with pd.ExcelWriter(median_touchtime_per_session_per_mouse, engine='xlsxwriter') as writer:
    result.to_excel(writer)
    
#%% step types per muis per session
# this part of the code calculates the steptypes per mouse per session from the Steps_front 
# excel file and writes the results into a new excel file
ercc1_mut = pd.read_excel(steps_front_file_mut)
ercc1_ctr = pd.read_excel(steps_front_file_ctr)


df = pd.DataFrame (ercc1_mut)
df2 = pd.DataFrame(ercc1_ctr)
dff2 = df2.groupby(["subject id","session nr"]).sum()
dff2.pop("session id")
dff2.pop("run id")
dff2.pop("direction")
dff = df.groupby(["subject id","session nr"]).sum()
dff.pop("session id")
dff.pop("run id")
dff.pop("direction")
with pd.ExcelWriter(step_type_per_mouse_per_session, engine='xlsxwriter') as writer:
    dff.to_excel(writer, sheet_name='MUT')
    dff2.to_excel(writer,sheet_name='CTR')
    
#%% all runtimes
# this part of the code makes a figure displaying the distribution of the runtimes 
# at 3 time intervals for both the KO and the WT group 
y1=[]
y2=[]
y3=[]
y4=[]
y5=[]
y6=[]
ercc1_mut = pd.read_excel(parameter_file_mut)
ercc1_ctr = pd.read_excel(parameter_file_ctr)

# delete runs outside runtime window
ercc1_mut = ercc1_mut.drop(ercc1_mut[ercc1_mut.timerun < 800].index)
ercc1_mut = ercc1_mut.drop(ercc1_mut[ercc1_mut.timerun > 15000].index)
ercc1_ctr = ercc1_ctr.drop(ercc1_ctr[ercc1_ctr.timerun > 15000].index)
ercc1_ctr = ercc1_ctr.drop(ercc1_ctr[ercc1_ctr.timerun < 800].index)

#create the 3 time points at which we want to see the data
ko_1 = ercc1_mut[ercc1_mut["session nr"] == first_session_number]
ko_2 = ercc1_mut[ercc1_mut["session nr"] == middle_session_number]
ko_3 = ercc1_mut[ercc1_mut["session nr"] == last_session_number]
wt_1 = ercc1_ctr[ercc1_ctr["session nr"] == first_session_number]
wt_2 = ercc1_ctr[ercc1_ctr["session nr"] == middle_session_number]
wt_3 = ercc1_ctr[ercc1_ctr["session nr"] == last_session_number]

y1.extend(ko_1["timerun"])
y2.extend(ko_2["timerun"])
y3.extend(ko_3["timerun"])
y4.extend(wt_1["timerun"])
y5.extend(wt_2["timerun"])
y6.extend(wt_3["timerun"])

# create the plot, adjust the bw to own preference / optimal visual representation
# the higher the bw the smoother the curve gets.
plt.figure()
bw=0.35

sns.distplot(y1,label=('knockout first session'), hist = False, kde = True, kde_kws = {'bw': bw})
sns.distplot(y2,label=('knockout middle session'), hist = False, kde = True, kde_kws = {'bw': bw})
sns.distplot(y3,label=('knockout last session'), hist = False, kde = True, kde_kws = {'bw': bw})
sns.distplot(y4,label=('wildtype first session'), hist = False, kde = True, kde_kws = {'bw': bw})
sns.distplot(y5,label=('wildtype middle session'), hist = False, kde = True, kde_kws = {'bw': bw})
sns.distplot(y6,label=('wildtype last session'), hist = False, kde = True, kde_kws = {'bw': bw})
plt.legend()
plt.xlabel('runtime')
plt.xlim(0,8000)
plt.savefig(runtime)

#%% all touchtimes
# this part of the code makes a figure displaying the distribution of the runtimes 
# at 3 time intervals for both the KO and the WT group 
y1=[]
y2=[]
y3=[]
y4=[]
y5=[]
y6=[]
ercc1_mut = pd.read_excel(alltouches_file_mut)
ercc1_ctr = pd.read_excel(alltouches_file_ctr)

ko_1 = ercc1_mut[ercc1_mut["sessionnr"] == first_session_number]
ko_2 = ercc1_mut[ercc1_mut["sessionnr"] == middle_session_number]
ko_3 = ercc1_mut[ercc1_mut["sessionnr"] == last_session_number]
wt_1 = ercc1_ctr[ercc1_ctr["sessionnr"] == first_session_number]
wt_2 = ercc1_ctr[ercc1_ctr["sessionnr"] == middle_session_number]
wt_3 = ercc1_ctr[ercc1_ctr["sessionnr"] == last_session_number]

y1.extend(ko_1["touch"])
y2.extend(ko_2["touch"])
y3.extend(ko_3["touch"])
y4.extend(wt_1["touch"])
y5.extend(wt_2["touch"])
y6.extend(wt_3["touch"])


df1 = pd.DataFrame(y1)
df2 = pd.DataFrame(y2)
df3 = pd.DataFrame(y3)
df4 = pd.DataFrame(y4)
df5 = pd.DataFrame(y5)
df6 = pd.DataFrame(y6)
df_final =[df1,df2,df3,df4,df5,df6]
test= pd.concat(df_final,axis=1)
test.columns = ["ko_first_session", "ko_middle_session","ko_last_session","wt_first_session","wt_middle_session","wt_last_session"]

cut_off_time = 500

y11 = [item for item in y1 if item <= cut_off_time]
y21 = [item for item in y2 if item <= cut_off_time]
y31 = [item for item in y3 if item <= cut_off_time]
y41 = [item for item in y4 if item <= cut_off_time]
y51 = [item for item in y5 if item <= cut_off_time]
y61 = [item for item in y6 if item <= cut_off_time]


bw=0.12
plt.figure()
sns.distplot(y11,label=('knockout first session'),  hist = False, kde = True, kde_kws = {'bw' : bw})
sns.distplot(y21,label=('knockout middle session'), hist = False, kde = True, kde_kws = {'bw' : bw})
sns.distplot(y31,label=('knockout last session'), hist = False, kde = True, kde_kws = {'bw' : bw})
sns.distplot(y41,label=('wildtype first session'), hist = False, kde = True, kde_kws = {'bw' : bw})
sns.distplot(y51,label=('wildtype middle session'), hist = False, kde = True, kde_kws = {'bw' : bw})
sns.distplot(y61,label=('wildtype last session'),hist = False, kde = True, kde_kws = {'bw' : bw})
plt.xlabel('touch time')
plt.ylabel('amount of touches')
plt.xlim(0,500)
plt.legend()
plt.savefig(touchtime_alltouches)


#%% runtypes 
# this part of the code calculates the runtypes for the KO and WT group from the parameter excel file.

df = pd.read_excel(parameter_file_mut)
light=[]
air=[]
returns=[]
escape=[]
irregular=[]
irregularses=[]
airses=[]
escapeses=[]
returnsses=[]
lightses=[]

# Loop through the rows of the dataframe and determine what run falls under which category.
for index, row in df.iterrows():
  if ('4-6-4-6' in row['state sequence']) & ('1-2-1' in row ["state sequence"]):
    irregular.append(row['subject id'])
    irregularses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '1-2-1' in row['state sequence']:
    escape.append(row['subject id'])
    escapeses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '4-6-4-6' in row['state sequence']:
    returns.append(row['subject id'])
    returnsses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '1-3-6' in row['state sequence']:
    light.append(row['subject id'])
    lightses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '1-3-4-6' in row['state sequence']:
    air.append(row['subject id'])
    airses.append(row['session nr'])
    df.drop(index, inplace=True)



escape = pd.DataFrame(escape)
returns = pd.DataFrame(returns)
light = pd.DataFrame(light)
air = pd.DataFrame(air)
irregular = pd.DataFrame(irregular)
escapeses = pd.DataFrame(escapeses)
airses = pd.DataFrame(airses)
lightses = pd.DataFrame(lightses)
returnsses = pd.DataFrame(returnsses)
irregularses = pd.DataFrame(irregularses)

df_final = [escape,escapeses]
df_final2= [returns,returnsses]
df_final3 = [air,airses]
df_final4 = [light,lightses]
df_final5 =[irregular,irregularses]
df_final6 = [escape,escapeses,returns,returnsses,air,airses,light,lightses,irregular,irregularses]
test=pd.concat(df_final,axis=1)
test2 = pd.concat(df_final2,axis=1)
test3 = pd.concat(df_final3,axis=1)
test4 = pd.concat(df_final4,axis=1)
test5 = pd.concat(df_final5,axis=1)
test6 = pd.concat(df_final6, axis=1)
test.columns = ["subjectid","session nr"]
test2.columns=["subjectid","session nr"]
test3.columns=["subjectid","session nr"]
test4.columns=["subjectid","session nr"]
test5.columns=["subjectid","session nr"]
test=test.groupby(["subjectid","session nr"]).size()
test2=test2.groupby(["subjectid","session nr"]).size()
test3=test3.groupby(["subjectid","session nr"]).size()
test4=test4.groupby(["subjectid","session nr"]).size()
test5=test5.groupby(["subjectid","session nr"]).size()

final_test = [test3,test2,test,test4,test5]
final_test=pd.concat(final_test,axis=1)
final_test.columns = ["air","return","escape","light","irregular"]
with pd.ExcelWriter(runtypes_mut, engine='xlsxwriter') as writer:
      final_test.to_excel(writer, sheet_name = 'all')
df= pd.read_excel(parameter_file_ctr)
light=[]
air=[]
returns=[]
escape=[]
irregular=[]
irregularses=[]
airses=[]
escapeses=[]
returnsses=[]
lightses=[]

# Loop through the rows of the dataframe and see for each run what type is should be . 
for index, row in df.iterrows():
  if ('4-6-4-6' in row['state sequence']) & ('1-2-1' in row ["state sequence"]):
    irregular.append(row['subject id'])
    irregularses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '1-2-1' in row['state sequence']:
    escape.append(row['subject id'])
    escapeses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '4-6-4-6' in row['state sequence']:
    returns.append(row['subject id'])
    returnsses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '1-3-6' in row['state sequence']:
    light.append(row['subject id'])
    lightses.append(row['session nr'])
    df.drop(index, inplace=True)
  elif '1-3-4-6' in row['state sequence']:
    air.append(row['subject id'])
    airses.append(row['session nr'])
    df.drop(index, inplace=True)



escape = pd.DataFrame(escape)
returns = pd.DataFrame(returns)
light = pd.DataFrame(light)
air = pd.DataFrame(air)
irregular = pd.DataFrame(irregular)
escapeses = pd.DataFrame(escapeses)
airses = pd.DataFrame(airses)
lightses = pd.DataFrame(lightses)
returnsses = pd.DataFrame(returnsses)
irregularses = pd.DataFrame(irregularses)

df_final = [escape,escapeses]
df_final2= [returns,returnsses]
df_final3 = [air,airses]
df_final4 = [light,lightses]
df_final5 =[irregular,irregularses]
df_final6 = [escape,escapeses,returns,returnsses,air,airses,light,lightses,irregular,irregularses]
test=pd.concat(df_final,axis=1)
test2 = pd.concat(df_final2,axis=1)
test3 = pd.concat(df_final3,axis=1)
test4 = pd.concat(df_final4,axis=1)
test5 = pd.concat(df_final5,axis=1)
test6 = pd.concat(df_final6, axis=1)
test.columns = ["subjectid","session nr"]
test2.columns=["subjectid","session nr"]
test3.columns=["subjectid","session nr"]
test4.columns=["subjectid","session nr"]
test5.columns=["subjectid","session nr"]
test=test.groupby(["subjectid","session nr"]).size()
test2=test2.groupby(["subjectid","session nr"]).size()
test3=test3.groupby(["subjectid","session nr"]).size()
test4=test4.groupby(["subjectid","session nr"]).size()
test5=test5.groupby(["subjectid","session nr"]).size()

final_test = [test3,test2,test,test4,test5]
final_test=pd.concat(final_test,axis=1)
final_test.columns = ["air","return","escape","light","irregular"]
with pd.ExcelWriter(runtypes_ctr, engine='xlsxwriter') as writer:
      final_test.to_excel(writer, sheet_name = 'all')
#%% duty cycle / cv2  per muis per sessie 
# this part of the code calculates the Duty cycle and the CV2 for each individual paw per mouse per session. 
swingtimeRF=[]
swingtime=[]
swingtimeRH=[]
swingtimeLH=[]

df = pd.read_excel(alltouches_file_mut)
df = df.iloc[: , 1:]

dfLH = df
dfRF = df
dfRH = df

df = df[df["paw"].str.contains("rh") == False]
df = df[df["paw"].str.contains("lh") == False]
df = df[df["paw"].str.contains("rf") == False]

dfLH = dfLH[dfLH["paw"].str.contains("rh") == False]
dfLH = dfLH[dfLH["paw"].str.contains("lf") == False]
dfLH = dfLH[dfLH["paw"].str.contains("rf") == False]

dfRH = dfRH[dfRH["paw"].str.contains("lh") == False]
dfRH = dfRH[dfRH["paw"].str.contains("lf") == False]
dfRH = dfRH[dfRH["paw"].str.contains("rf") == False]

dfRF = dfRF[dfRF["paw"].str.contains("lh") == False]
dfRF = dfRF[dfRF["paw"].str.contains("lf") == False]
dfRF = dfRF[dfRF["paw"].str.contains("rh") == False]

df["diff"] = df["run id"].diff()
dfLH["diff"] = dfLH["run id"].diff()
dfRH["diff"] = dfRH["run id"].diff()
dfRF["diff"] = dfRF["run id"].diff()

    
with pd.ExcelWriter(test_file, engine='xlsxwriter') as writer:
      df.to_excel(writer, sheet_name = 'LF')
      dfLH.to_excel(writer, sheet_name = 'LH')
      dfRF.to_excel(writer, sheet_name = 'RF')
      dfRH.to_excel(writer, sheet_name = 'RH')


df2 = pd.read_excel(test_file, sheet_name='LF', header = 0, index_col=None)
df2 = df2.iloc[: , 1:]
df2LH = pd.read_excel(test_file, sheet_name='LH', header = 0, index_col=None)
df2LH = df2LH.iloc[: , 1:]
df2RH = pd.read_excel(test_file, sheet_name='RH', header = 0, index_col=None)
df2RH = df2RH.iloc[: , 1:]
df2RF = pd.read_excel(test_file, sheet_name='RF', header = 0, index_col=None)
df2RF = df2RF.iloc[: , 1:]

index_list = df2.index[df2["diff"] != 0].tolist()
index_listLH = df2LH.index[df2LH["diff"] != 0].tolist()
index_listRH = df2RH.index[df2RH["diff"] != 0].tolist()
index_listRF = df2RF.index[df2RF["diff"] != 0].tolist()


def subtract_from_columns(df2, index_list, value_column):
  for i in range(len(index_list) - 1):
    start_index = index_list[i]
    end_index = index_list[i+1]
    value = df2.at[start_index, value_column]
    df2.iloc[start_index:end_index, 8] -= value
    df2.iloc[start_index:end_index, 9] -= value
  return df2

value_column = 'begintouches'

result = subtract_from_columns(df2,index_list,value_column) 
for i in range(len(result)-1):
    if result["direction"].iloc[i] == 0:
        result["rung"][i] = (result["rung"][i] - 38)*-1
    swingtime.append(result["begintouches"].iloc[i+1]-result["endtouches"].iloc[i])
swingtime.insert(0,0)
result["swingtime"] = swingtime
mask = result['diff'] != 0
result.loc[mask, 'swingtime'] = pd.np.nan
            
resultLH = subtract_from_columns(df2LH,index_listLH,value_column) 
for i in range(len(resultLH)-1):
    if resultLH["direction"].iloc[i] == 0:
        resultLH["rung"][i] = (resultLH["rung"][i] - 38)*-1
    swingtimeLH.append(resultLH["begintouches"].iloc[i+1]-resultLH["endtouches"].iloc[i])
swingtimeLH.insert(0,0)   
resultLH["swingtime"] = swingtimeLH 
mask = resultLH['diff'] != 0
resultLH.loc[mask, 'swingtime'] = pd.np.nan
       
resultRH = subtract_from_columns(df2RH,index_listRH,value_column) 
for i in range(len(resultRH)-1):
    if resultRH["direction"].iloc[i] == 0:
        resultRH["rung"][i] = (resultRH["rung"][i] - 38)*-1
    swingtimeRH.append(resultRH["begintouches"].iloc[i+1]-resultRH["endtouches"].iloc[i])
swingtimeRH.insert(0,0)   
resultRH["swingtime"] = swingtimeRH
mask = resultRH['diff'] != 0
resultRH.loc[mask, 'swingtime'] = pd.np.nan
  
resultRF = subtract_from_columns(df2RF,index_listRF,value_column) 
for i in range(len(resultRF)-1):
    if resultRF["direction"].iloc[i] == 0:
        resultRF["rung"][i] = (resultRF["rung"][i] - 38)*-1
    swingtimeRF.append(resultRF["begintouches"].iloc[i+1]-resultRF["endtouches"].iloc[i])
swingtimeRF.insert(0,0)    
resultRF["swingtime"] = swingtimeRF
mask = resultRF['diff'] != 0
resultRF.loc[mask, 'swingtime'] = pd.np.nan


merge = [result,resultRF,resultRH,resultLH]
merger=pd.concat(merge,axis=0)
merger = merger.reset_index(drop=True)

indexes = []
indexes2= []
for index, row in merger.iterrows():
  if row['diff'] != 0:
    indexes2.append(index)
    
for index, row in merger.iterrows():
  if row['swingtime'] < 0:
    indexes.append(index)


with pd.ExcelWriter(MUT_inbetween_file, engine='xlsxwriter') as writer:
      result.to_excel(writer, sheet_name = 'LF')
      resultLH.to_excel(writer, sheet_name = 'LH')
      resultRH.to_excel(writer, sheet_name = 'RH')
      resultRF.to_excel(writer, sheet_name = 'RF')
      merger.to_excel(writer, sheet_name = 'merger')

df3 = pd.read_excel(MUT_inbetween_file, sheet_name='merger', header = 0, index_col=None)
steptime = []
for i in range(len(df3)-1):
    steptime.append(df3["touch"].iloc[i]+df3["swingtime"].iloc[i+1])
steptime.insert(0,0)    
df3["steptime"]=steptime
df3 = df3.iloc[: , 1:]

df3.reset_index()


with pd.ExcelWriter(MUT_inbetween_file, engine='xlsxwriter') as writer:
      result.to_excel(writer, sheet_name = 'LF')
      resultLH.to_excel(writer, sheet_name = 'LH')
      resultRH.to_excel(writer, sheet_name = 'RH')
      resultRF.to_excel(writer, sheet_name = 'RF')
      merger.to_excel(writer, sheet_name = 'merger')
      df3.to_excel(writer, sheet_name = 'merger+steptime')

df4 = pd.read_excel(MUT_inbetween_file, sheet_name= 'merger+steptime')
df2 = df4[['diff', 'swingtime','steptime']]
df4 = df4.drop(['diff', 'swingtime','steptime'], axis=1)

df2 =df2[1:]
df2 = df2.append(pd.Series([np.nan]*len(df2.columns), index=df2.columns), ignore_index=True)
df4 = df4.iloc[: , 1:]
df_concatMUT = pd.concat([df4, df2], axis=1)

df_concatMUT['DC'] = df_concatMUT['touch']/df_concatMUT['steptime']
cv2=[]
for i in range(len(df_concatMUT)-1):
    cv2.append(np.abs(2*(df_concatMUT["steptime"].iloc[i+1]-df_concatMUT["steptime"].iloc[i])/(df_concatMUT["steptime"].iloc[i+1]+df_concatMUT["steptime"].iloc[i])))
cv2.insert(-1,0)    
df_concatMUT["cv2"]=cv2

swingtimeRF=[]
swingtime=[]
swingtimeRH=[]
swingtimeLH=[]
df = pd.read_excel(alltouches_file_ctr)
df = df.iloc[: , 1:]

dfLH = df
dfRF = df
dfRH = df

df = df[df["paw"].str.contains("rh") == False]
df = df[df["paw"].str.contains("lh") == False]
df = df[df["paw"].str.contains("rf") == False]

dfLH = dfLH[dfLH["paw"].str.contains("rh") == False]
dfLH = dfLH[dfLH["paw"].str.contains("lf") == False]
dfLH = dfLH[dfLH["paw"].str.contains("rf") == False]

dfRH = dfRH[dfRH["paw"].str.contains("lh") == False]
dfRH = dfRH[dfRH["paw"].str.contains("lf") == False]
dfRH = dfRH[dfRH["paw"].str.contains("rf") == False]

dfRF = dfRF[dfRF["paw"].str.contains("lh") == False]
dfRF = dfRF[dfRF["paw"].str.contains("lf") == False]
dfRF = dfRF[dfRF["paw"].str.contains("rh") == False]

df["diff"] = df["run id"].diff()
dfLH["diff"] = dfLH["run id"].diff()
dfRH["diff"] = dfRH["run id"].diff()
dfRF["diff"] = dfRF["run id"].diff()

    
with pd.ExcelWriter(test_file, engine='xlsxwriter') as writer:
      df.to_excel(writer, sheet_name = 'LF')
      dfLH.to_excel(writer, sheet_name = 'LH')
      dfRF.to_excel(writer, sheet_name = 'RF')
      dfRH.to_excel(writer, sheet_name = 'RH')


df2 = pd.read_excel(test_file, sheet_name='LF', header = 0, index_col=None)
df2 = df2.iloc[: , 1:]
df2LH = pd.read_excel(test_file, sheet_name='LH', header = 0, index_col=None)
df2LH = df2LH.iloc[: , 1:]
df2RH = pd.read_excel(test_file, sheet_name='RH', header = 0, index_col=None)
df2RH = df2RH.iloc[: , 1:]
df2RF = pd.read_excel(test_file, sheet_name='RF', header = 0, index_col=None)
df2RF = df2RF.iloc[: , 1:]

index_list = df2.index[df2["diff"] != 0].tolist()
index_listLH = df2LH.index[df2LH["diff"] != 0].tolist()
index_listRH = df2RH.index[df2RH["diff"] != 0].tolist()
index_listRF = df2RF.index[df2RF["diff"] != 0].tolist()

result = subtract_from_columns(df2,index_list,value_column) 
for i in range(len(result)-1):
    if result["direction"].iloc[i] == 0:
        result["rung"][i] = (result["rung"][i] - 38)*-1
    swingtime.append(result["begintouches"].iloc[i+1]-result["endtouches"].iloc[i])
swingtime.insert(0,0)
result["swingtime"] = swingtime
mask = result['diff'] != 0
result.loc[mask, 'swingtime'] = pd.np.nan
            
resultLH = subtract_from_columns(df2LH,index_listLH,value_column) 
for i in range(len(resultLH)-1):
    if resultLH["direction"].iloc[i] == 0:
        resultLH["rung"][i] = (resultLH["rung"][i] - 38)*-1
    swingtimeLH.append(resultLH["begintouches"].iloc[i+1]-resultLH["endtouches"].iloc[i])
swingtimeLH.insert(0,0)   
resultLH["swingtime"] = swingtimeLH 
mask = resultLH['diff'] != 0
resultLH.loc[mask, 'swingtime'] = pd.np.nan
       
resultRH = subtract_from_columns(df2RH,index_listRH,value_column) 
for i in range(len(resultRH)-1):
    if resultRH["direction"].iloc[i] == 0:
        resultRH["rung"][i] = (resultRH["rung"][i] - 38)*-1
    swingtimeRH.append(resultRH["begintouches"].iloc[i+1]-resultRH["endtouches"].iloc[i])
swingtimeRH.insert(0,0)   
resultRH["swingtime"] = swingtimeRH
mask = resultRH['diff'] != 0
resultRH.loc[mask, 'swingtime'] = pd.np.nan
  
resultRF = subtract_from_columns(df2RF,index_listRF,value_column) 
for i in range(len(resultRF)-1):
    if resultRF["direction"].iloc[i] == 0:
        resultRF["rung"][i] = (resultRF["rung"][i] - 38)*-1
    swingtimeRF.append(resultRF["begintouches"].iloc[i+1]-resultRF["endtouches"].iloc[i])
swingtimeRF.insert(0,0)    
resultRF["swingtime"] = swingtimeRF
mask = resultRF['diff'] != 0
resultRF.loc[mask, 'swingtime'] = pd.np.nan


merge = [result,resultRF,resultRH,resultLH]
merger=pd.concat(merge,axis=0)
merger = merger.reset_index(drop=True)

indexes = []
indexes2= []
for index, row in merger.iterrows():
  if row['diff'] != 0:
    indexes2.append(index)
    
for index, row in merger.iterrows():
  if row['swingtime'] < 0:
    indexes.append(index)

with pd.ExcelWriter(CTR_inbetween_file, engine='xlsxwriter') as writer:
      result.to_excel(writer, sheet_name = 'LF')
      resultLH.to_excel(writer, sheet_name = 'LH')
      resultRH.to_excel(writer, sheet_name = 'RH')
      resultRF.to_excel(writer, sheet_name = 'RF')
      merger.to_excel(writer, sheet_name = 'merger')

df3 = pd.read_excel(CTR_inbetween_file, sheet_name='merger', header = 0, index_col=None)
steptime = []
for i in range(len(df3)-1):
    steptime.append(df3["touch"].iloc[i]+df3["swingtime"].iloc[i+1])
steptime.insert(0,0)    
df3["steptime"]=steptime
df3 = df3.iloc[: , 1:]

df3.reset_index()


with pd.ExcelWriter(CTR_inbetween_file, engine='xlsxwriter') as writer:
      result.to_excel(writer, sheet_name = 'LF')
      resultLH.to_excel(writer, sheet_name = 'LH')
      resultRH.to_excel(writer, sheet_name = 'RH')
      resultRF.to_excel(writer, sheet_name = 'RF')
      merger.to_excel(writer, sheet_name = 'merger')
      df3.to_excel(writer, sheet_name = 'merger+steptime')

df4 = pd.read_excel(CTR_inbetween_file, sheet_name= 'merger+steptime')
df2 = df4[['diff', 'swingtime','steptime']]
df4 = df4.drop(['diff', 'swingtime','steptime'], axis=1)

df2 =df2[1:]
df2 = df2.append(pd.Series([np.nan]*len(df2.columns), index=df2.columns), ignore_index=True)
df4 = df4.iloc[: , 1:]
df_concatCTR = pd.concat([df4, df2], axis=1)

df_concatCTR['DC'] = df_concatCTR['touch']/df_concatCTR['steptime']
cv2=[]
for i in range(len(df_concatCTR)-1):
    cv2.append(np.abs(2*(df_concatCTR["steptime"].iloc[i+1]-df_concatCTR["steptime"].iloc[i])/(df_concatCTR["steptime"].iloc[i+1]+df_concatCTR["steptime"].iloc[i])))
cv2.insert(-1,0)    
df_concatCTR["cv2"]=cv2

df = df_concatMUT
dff=df_concatCTR

dfLF = df[df['paw'] == 'lf']
dfRF = df[df['paw'] == 'rf']
dfRH = df[df['paw'] == 'rh']
dfLH = df[df['paw'] == 'lh']


dfLF = dfLF.groupby(["subject id","sessionnr"]).median()
dfLH = dfLH.groupby(["subject id","sessionnr"]).median()
dfRF = dfRF.groupby(["subject id","sessionnr"]).median()
dfRH = dfRH.groupby(["subject id","sessionnr"]).median()

columns_to_remove = ["touch", "rung", "begintouches", "endtouches", "sides", "direction", "run id", "diff", "swingtime"]
dfLF = dfLF.drop(columns=columns_to_remove)
dfRF = dfRF.drop(columns=columns_to_remove)
dfRH = dfRH.drop(columns=columns_to_remove)
dfLH = dfLH.drop(columns=columns_to_remove)


test = [dfLF,dfRF,dfLH,dfRH]
final=pd.concat(test,axis=1)
final.columns =["steptimeLF","DCLF","CV2LF","steptimeRF","DCRF","CV2RF","steptimeLH","DCLH","CV2LH","steptimeRH","DCRH","CV2RH"]


dffLF = dff[dff['paw'] == 'lf']
dffRF = dff[dff['paw'] == 'rf']
dffRH = dff[dff['paw'] == 'rh']
dffLH = dff[dff['paw'] == 'lh']


dffLF = dffLF.groupby(["subject id","sessionnr"]).median()
dffLH = dffLH.groupby(["subject id","sessionnr"]).median()
dffRF = dffRF.groupby(["subject id","sessionnr"]).median()
dffRH = dffRH.groupby(["subject id","sessionnr"]).median()

dffLF = dffLF.drop(columns=columns_to_remove)
dffRF = dffRF.drop(columns=columns_to_remove)
dffRH = dffRH.drop(columns=columns_to_remove)
dffLH = dffLH.drop(columns=columns_to_remove)


test2 = [dffLF,dffRF,dffLH,dffRH]
final2=pd.concat(test2,axis=1)
final2.columns =["steptimeLF","DCLF","CV2LF","steptimeRF","DCRF","Cv2RF","steptimeLH","DCLH","CV2LH","steptimeRH","DCRH","Cv2RH"]


with pd.ExcelWriter(median_DC_CV2_Steptime, engine='xlsxwriter') as writer:
    final.to_excel(writer, sheet_name = 'MUT')
    final2.to_excel(writer, sheet_name = 'CTR')
#%% PCA this is for 2 groups for the PCA with 4 groups the other 2 groups have to be added manually
df = df_concatCTR
df = df[df["paw"].str.contains("rh") == False]
df = df[df["paw"].str.contains("lh") == False]
df = df[df["paw"].str.contains("lf") == False]

columns_to_include = ['run id', 'touch','sessionnr']
touch_df = df[columns_to_include]

columns_to_include = ['run id', 'steptime','sessionnr']
steptime_df = df[columns_to_include]
steptime_df = steptime_df.dropna(subset=["steptime"])


columns_to_include = ['run id', 'cv2','sessionnr']
cv2_df = df[columns_to_include]
cv2_df = cv2_df.dropna(subset=["cv2"])

columns_to_include = ['run id', 'DC','sessionnr']
DC_df = df[columns_to_include]
DC_df = DC_df.dropna(subset=["DC"])

touch_df = touch_df.groupby(["run id"]).median()
cv2_df = cv2_df.groupby(["run id"]).median()
DC_df = DC_df.groupby(["run id"]).median()
steptime_df = steptime_df.groupby(["run id"]).median()

touch_df = touch_df.merge(cv2_df, how='inner', on='run id').copy()
DC_df = DC_df.merge(cv2_df, how='inner', on='run id').copy()
steptime_df = steptime_df.merge(DC_df, how='inner', on='run id').copy()

df2 = pd.read_excel(parameter_file_ctr)
df2 = df2.drop_duplicates()
df2 = df2.merge(steptime_df, how='inner', on='run id').copy()
df2 = df2.drop(df2.columns[[36,38,40]],axis=1)


df2 = df2.drop(columns=['swing phase rf','swing phase lf','swing phase lh','odd steps','swing phase rh','direction','support 0 paws','state sequence'])
with pd.ExcelWriter(Pca_prep_ctr, engine='xlsxwriter') as writer:
      df2.to_excel(writer)
df = df_concatMUT
df = df[df["paw"].str.contains("rh") == False]
df = df[df["paw"].str.contains("lh") == False]
df = df[df["paw"].str.contains("lf") == False]

columns_to_include = ['run id', 'touch','sessionnr']
touch_df = df[columns_to_include]

columns_to_include = ['run id', 'steptime','sessionnr']
steptime_df = df[columns_to_include]
steptime_df = steptime_df.dropna(subset=["steptime"])


columns_to_include = ['run id', 'cv2','sessionnr']
cv2_df = df[columns_to_include]
cv2_df = cv2_df.dropna(subset=["cv2"])

columns_to_include = ['run id', 'DC','sessionnr']
DC_df = df[columns_to_include]
DC_df = DC_df.dropna(subset=["DC"])

touch_df = touch_df.groupby(["run id"]).median()
cv2_df = cv2_df.groupby(["run id"]).median()
DC_df = DC_df.groupby(["run id"]).median()
steptime_df = steptime_df.groupby(["run id"]).median()

touch_df = touch_df.merge(cv2_df, how='inner', on='run id').copy()
DC_df = DC_df.merge(cv2_df, how='inner', on='run id').copy()
steptime_df = steptime_df.merge(DC_df, how='inner', on='run id').copy()

df2 = pd.read_excel(parameter_file_mut)
df2 = df2.drop_duplicates()
df2 = df2.merge(steptime_df, how='inner', on='run id').copy()
df2 = df2.drop(df2.columns[[36,38,40]],axis=1)


df2 = df2.drop(columns=['swing phase rf','swing phase lf','swing phase lh','odd steps','swing phase rh','direction','support 0 paws','state sequence'])
with pd.ExcelWriter(Pca_prep_mut, engine='xlsxwriter') as writer:
      df2.to_excel(writer)

ercc1_mut = pd.read_excel(Pca_prep_mut)
ercc1_ctr = pd.read_excel(Pca_prep_ctr)
min_session_nr = ercc1_mut["session nr"].min()
max_session_nr = ercc1_mut["session nr"].max()

# Create a dictionary to hold the Eko dataframes
eko_dict = {}

# Loop over the session numbers and create the Eko dataframes
for session_nr in range(min_session_nr, max_session_nr+1):
    eko_name = f"Eko_{session_nr}"
    eko_dict[eko_name] = ercc1_mut[ercc1_mut["session nr"] == session_nr]
for key, value in eko_dict.items():
    value[key] = key
    value.rename(columns={key: "genotype"}, inplace=True)

# add column to each group with the label before adding them together making one big dataframe

groups = pd.concat(eko_dict.values(), ignore_index=True)

features = groups[groups.columns[5:31]]

labels = groups["genotype"]
sessions = groups["session nr"]
subjects = groups["subject id"]
#normalize the data 
features_norm = StandardScaler().fit_transform(features)
features_norm = pd.DataFrame(features_norm, columns= features.columns)
# keep making components until 85% variance explained 
pca = PCA(n_components = 0.85)


principal_comp = pca.fit_transform(features_norm)
principal_df = pd.DataFrame(data = principal_comp)
final_df = pd.concat([principal_df, labels,sessions,subjects], axis = 1)

final_df=final_df.groupby(["genotype"]).median().reset_index()
with pd.ExcelWriter(scatterplot_pca, engine='xlsxwriter') as writer:
    final_df.to_excel(writer)
pca.explained_variance_ratio_
# scree plot
pc_values = np.arange(pca.n_components_) + 1
plt.figure()
plt.plot(pc_values, pca.explained_variance_ratio_, 'ro-')
plt.savefig(EV_ratio)
plt.show()

print ("Proportion of Variance Explained : ", pca.explained_variance_ratio_)    
out_sum = np.cumsum(pca.explained_variance_ratio_)  
print ("Cumulative Prop. Variance Explained: ", out_sum)
# plot pca
fig = plt.figure(figsize = (8,8))
ax = fig.add_subplot(1,1,1)
ax.set_xlabel('Principal Component 1', fontsize = 15)
ax.set_ylabel('Principal Component 2', fontsize = 15)
ax.set_title('2 component PCA', fontsize = 20)

eko_names = list(eko_dict.keys())

targets = eko_names

colors = ['#F98B88','#F75D59','#FF6347','#FF0000','#F70D1A','#F62817','#16E2F5','#0AFFFF','#57FEFF','#9AFEFF','#AFDCEC','#82CAFF']


for target, color in zip(targets, colors):
    indicesToKeep = final_df["genotype"] == target
    points = ax.scatter(final_df.loc[indicesToKeep, 0], final_df.loc[indicesToKeep, 1], c = color, s = 50)
ax.legend(targets, loc='lower right',fontsize=6)
ax.grid()
plt.savefig(PCA_plot, format="eps")
plt.close()
components = pca.components_
# Create a DataFrame from the components array
df = pd.DataFrame(components, columns=['time run','N.o.Steps','Even steps','step2','step4','step6','stance phase lf','stance phase lh','stance phase rf','stance phase rh','lag lf/rh','lag rf/lh','support diagonal','support girlde','support lateral','support 1 paw','support 3 paws','support 4paws','high rungs front','high rungs hind','low rungs front','low rungs hind','mean touchtime','steptime','DC','cv2'])

# Write the DataFrame to an Excel file
df.to_excel(feature_per_component_pca)

# features_train, features_test, labels_train, labels_test = train_test_split(features_norm, labels, test_size = 0.25, stratify = labels)

# # create model
# forest = RandomForestClassifier(n_estimators = 20000, max_features = 'auto', max_samples= None)

# # train model
# forest.fit(features_train, labels_train)

# # prediction
# pred = forest.predict(features_test)
# print("Accuracy:", metrics.accuracy_score(labels_test, pred))
with pd.ExcelWriter(PCA_values, engine='xlsxwriter') as writer:
    final_df.to_excel(writer)
#%% heatmaps The heatmaps that are made with the code are OF 1 PAIR OF PAWS 
#if you want all 3 pairs run this part of the code multiple times 
# and alter the variables to get different combinations of paws
swingtime=[]
df = pd.read_excel(alltouches_file_ctr)
df = df.iloc[: , 1:]

#select the paw pair we want
df = df[df["paw"].str.contains("rh") == False]
df = df[df["paw"].str.contains("lf") == False]
df2=df

#make sure that we dont calculate RF --> RF due to a missing / unclassified touch
for i in range(len(df)-1):
    if (df['paw'].iloc[i] == df['paw'].iloc[i+1]):
        df2 = df2.drop(df.index[i])
    else:
        i=i+1

# determine the diff column to seperate the runs 
df2["diff"] = df2["run id"].diff()


    
with pd.ExcelWriter(heatmaps_ctr_inbetween, engine='xlsxwriter') as writer:
      df2.to_excel(writer)



df2 = pd.read_excel(heatmaps_ctr_inbetween, header = 0, index_col=None)
df2 = df2.iloc[: , 1:]


index_list = df2.index[df2["diff"] != 0].tolist()

# set every run to start at time = 0 by using the diff column
result = subtract_from_columns(df2,index_list,value_column) 
for i in range(len(result)-2):
    if result["direction"].iloc[i] == 0:
        result["rung"][i] = (result["rung"][i] - 38)*-1
    swingtime.append(result["begintouches"].iloc[i+2]-result["endtouches"].iloc[i])
swingtime.insert(0,0)
swingtime.append(0)
result["swingtime"] = swingtime
mask = result['diff'] != 0
mask2 = result['swingtime']<=0
result.loc[mask, 'swingtime'] = pd.np.nan
result.loc[mask2, 'swingtime'] = pd.np.nan


merge = [result]
merger=pd.concat(merge,axis=0)
merger = merger.reset_index(drop=True)

indexes = []
indexes2= []
for index, row in merger.iterrows():
  if row['diff'] != 0:
    indexes2.append(index)
    
for index, row in merger.iterrows():
  if row['swingtime'] < 0:
    indexes.append(index)

with pd.ExcelWriter(heatmaps_ctr_RFLH, engine='xlsxwriter') as writer:
      merger.to_excel(writer)


df3 = pd.read_excel(heatmaps_ctr_RFLH, header = 0, index_col=None)
steptime = []

# add steptime column
for i in range(len(df3)-1):
    steptime.append(df3["touch"].iloc[i]+df3["swingtime"].iloc[i+1])
steptime.insert(0,0) 
df3["steptime"]=steptime
df3 = df3.iloc[: , 1:]

#add delay and duty cycle column
df3.reset_index()
delay = []
duty_cycle=[]
for i in range(len(df3)-1):
    delay.append((df3["begintouches"].iloc[i+1]-df3["begintouches"].iloc[i])/df3["steptime"].iloc[i+1])
    duty_cycle.append((df3["touch"].iloc[i])/df3["steptime"].iloc[i+1])
delay.insert(0,0) 
duty_cycle.insert(0,0) 
df3["delay"]=delay
df3["duty cycle"]=duty_cycle
df2 = df3[['delay', 'swingtime','steptime','duty cycle']]
df3 = df3.drop(['duty cycle','delay', 'swingtime','steptime'], axis=1)

df2 =df2[1:]
df2 = df2.append(pd.Series([np.nan]*len(df2.columns), index=df2.columns), ignore_index=True)
df3 = df3.iloc[: , 1:]
df_CTR = pd.concat([df3, df2], axis=1)

with pd.ExcelWriter(heatmaps_ctr_RFLH, engine='xlsxwriter') as writer:
      df_CTR.to_excel(writer)
     
# here we do the exact same but for the mutant mice group
swingtime=[]
df = pd.read_excel(alltouches_file_mut)
df = df.iloc[: , 1:]

df = df[df["paw"].str.contains("rh") == False]
df = df[df["paw"].str.contains("lf") == False]
df2=df
for i in range(len(df)-1):
    if (df['paw'].iloc[i] == df['paw'].iloc[i+1]):
        df2 = df2.drop(df.index[i])
    else:
        i=i+1

df2["diff"] = df2["run id"].diff()


    
with pd.ExcelWriter(heatmaps_mut_inbetween, engine='xlsxwriter') as writer:
      df2.to_excel(writer)



df2 = pd.read_excel(heatmaps_mut_inbetween, header = 0, index_col=None)
df2 = df2.iloc[: , 1:]


index_list = df2.index[df2["diff"] != 0].tolist()

result = subtract_from_columns(df2,index_list,value_column) 
for i in range(len(result)-2):
    if result["direction"].iloc[i] == 0:
        result["rung"][i] = (result["rung"][i] - 38)*-1
    swingtime.append(result["begintouches"].iloc[i+2]-result["endtouches"].iloc[i])
swingtime.insert(0,0)
swingtime.append(0)
result["swingtime"] = swingtime
mask = result['diff'] != 0
mask2 = result['swingtime']<=0
result.loc[mask, 'swingtime'] = pd.np.nan
result.loc[mask2, 'swingtime'] = pd.np.nan


merge = [result]
merger=pd.concat(merge,axis=0)
merger = merger.reset_index(drop=True)

indexes = []
indexes2= []
for index, row in merger.iterrows():
  if row['diff'] != 0:
    indexes2.append(index)
    
for index, row in merger.iterrows():
  if row['swingtime'] < 0:
    indexes.append(index)


with pd.ExcelWriter(heatmaps_mut_RFLH, engine='xlsxwriter') as writer:
      merger.to_excel(writer)



df3 = pd.read_excel(heatmaps_mut_RFLH)
steptime = []

for i in range(len(df3)-1):
    steptime.append(df3["touch"].iloc[i]+df3["swingtime"].iloc[i+1])
steptime.insert(0,0) 
df3["steptime"]=steptime
df3 = df3.iloc[: , 1:]

df3.reset_index()
delay = []
duty_cycle=[]
for i in range(len(df3)-1):
    delay.append((df3["begintouches"].iloc[i+1]-df3["begintouches"].iloc[i])/df3["steptime"].iloc[i+1])
    duty_cycle.append((df3["touch"].iloc[i])/df3["steptime"].iloc[i+1])
delay.insert(0,0) 
duty_cycle.insert(0,0) 
df3["delay"]=delay
df3["duty cycle"]=duty_cycle
df2 = df3[['delay', 'swingtime','steptime','duty cycle']]
df3 = df3.drop(['duty cycle','delay', 'swingtime','steptime'], axis=1)

df2 =df2[1:]
df2 = df2.append(pd.Series([np.nan]*len(df2.columns), index=df2.columns), ignore_index=True)
df3 = df3.iloc[: , 1:]
df_MUT = pd.concat([df3, df2], axis=1)

with pd.ExcelWriter(heatmaps_mut_RFLH, engine='xlsxwriter') as writer:
      df_MUT.to_excel(writer)

# filter the data to only get the delay of the RF paw compared to the LH paw not the other way around
# deletes delays smaller then 0 or bigger then 1 these can happen due to a step being unclassified and thus contain nonsense data
# also deletes values without delay this is the last touch a run
df = pd.read_excel(heatmaps_ctr_RFLH, header=0, index_col=None)
df = df[(df.delay >= 0) & (df.delay <= 1) & df["paw"].eq("rf")].dropna(subset=["delay"])
df = df.iloc[:,1:]

dfs = []
for i in range(first_session_number,first_session_number+amount_of_sessions):
    dfs.append(df[df["sessionnr"] == i])

df2 = pd.read_excel(heatmaps_mut_RFLH, header=0, index_col=None)
df2 = df2[(df2.delay >= 0) & (df2.delay <= 1) & df2["paw"].eq("rf")].dropna(subset=["delay"])
df2 = df2.iloc[:,1:]
for i in range(first_session_number,first_session_number+amount_of_sessions):
    dfs.append(df2[df2["sessionnr"] == i])

# Create a Pandas Excel writer using XlsxWriter as the engine
writer = pd.ExcelWriter(HBall_RFLH, engine='xlsxwriter')

# Iterate over the dataframes and write each to a separate sheet
for df in dfs:
    sheet_name = f"{df['subject id'].iloc[0]} - Session {df['sessionnr'].iloc[0]}"
    df.to_excel(writer, sheet_name=sheet_name, index=False)

# Save the Excel file
writer.save()
#%% making the heatmaps 

# loops through the excel sheets, every session has a sheet of data corresponding to this session.
# takes the proper columns of this sheets and plots these values in a 2d histogram from 0-->6.7%

import openpyxl
allvars = []
allsheetnames=[]
file = HBall_RFLH
workbook = openpyxl.load_workbook(file)
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows():
        data.append([cell.value for cell in row])
    df = pd.DataFrame(data)
    x = df[13]
    y = df[10]
    xlist = x.tolist()
    ylist = y.tolist()
    xlist = xlist[2:]
    ylist = ylist[2:]
    array=np.array([xlist,ylist])
    std_dev = np.round(np.std(array),4)
    allsheetnames.append(sheet_name)
    print(sheet_name)
    print(std_dev)
    scaling = len(xlist)*0.067
    xlist = [x * 100 for x in xlist]
    ylist = [y * 100 for y in ylist]
    plt.hist2d(xlist, ylist, bins = ([np.arange(0, 105, 5), np.arange(0, 105, 5)]), vmin=0, vmax=scaling, cmap = 'Greys')
    # plt.axis('off')
    plt.title(sheet_name)
    plt.colorbar()
    plt.savefig(heatmapsRFLH +'\\'+sheet_name+'', format='Tiff',bbox_inches='tight')
    plt.show()
#%% other paw configuration
swingtime=[]
df = pd.read_excel(alltouches_file_ctr)
df = df.iloc[: , 1:]

#select the paw pair we want
df = df[df["paw"].str.contains("rh") == False]
df = df[df["paw"].str.contains("lh") == False]
df2=df

#make sure that we dont calculate RF --> RF due to a missing / unclassified touch
for i in range(len(df)-1):
    if (df['paw'].iloc[i] == df['paw'].iloc[i+1]):
        df2 = df2.drop(df.index[i])
    else:
        i=i+1

# determine the diff column to seperate the runs 
df2["diff"] = df2["run id"].diff()


    
with pd.ExcelWriter(heatmaps_ctr_inbetween, engine='xlsxwriter') as writer:
      df2.to_excel(writer)



df2 = pd.read_excel(heatmaps_ctr_inbetween, header = 0, index_col=None)
df2 = df2.iloc[: , 1:]


index_list = df2.index[df2["diff"] != 0].tolist()

# set every run to start at time = 0 by using the diff column
result = subtract_from_columns(df2,index_list,value_column) 
for i in range(len(result)-2):
    if result["direction"].iloc[i] == 0:
        result["rung"][i] = (result["rung"][i] - 38)*-1
    swingtime.append(result["begintouches"].iloc[i+2]-result["endtouches"].iloc[i])
swingtime.insert(0,0)
swingtime.append(0)
result["swingtime"] = swingtime
mask = result['diff'] != 0
mask2 = result['swingtime']<=0
result.loc[mask, 'swingtime'] = pd.np.nan
result.loc[mask2, 'swingtime'] = pd.np.nan


merge = [result]
merger=pd.concat(merge,axis=0)
merger = merger.reset_index(drop=True)

indexes = []
indexes2= []
for index, row in merger.iterrows():
  if row['diff'] != 0:
    indexes2.append(index)
    
for index, row in merger.iterrows():
  if row['swingtime'] < 0:
    indexes.append(index)

with pd.ExcelWriter(heatmaps_ctr_RFLF, engine='xlsxwriter') as writer:
      merger.to_excel(writer)


df3 = pd.read_excel(heatmaps_ctr_RFLF, header = 0, index_col=None)
steptime = []

# add steptime column
for i in range(len(df3)-1):
    steptime.append(df3["touch"].iloc[i]+df3["swingtime"].iloc[i+1])
steptime.insert(0,0) 
df3["steptime"]=steptime
df3 = df3.iloc[: , 1:]

#add delay and duty cycle column
df3.reset_index()
delay = []
duty_cycle=[]
for i in range(len(df3)-1):
    delay.append((df3["begintouches"].iloc[i+1]-df3["begintouches"].iloc[i])/df3["steptime"].iloc[i+1])
    duty_cycle.append((df3["touch"].iloc[i])/df3["steptime"].iloc[i+1])
delay.insert(0,0) 
duty_cycle.insert(0,0) 
df3["delay"]=delay
df3["duty cycle"]=duty_cycle
df2 = df3[['delay', 'swingtime','steptime','duty cycle']]
df3 = df3.drop(['duty cycle','delay', 'swingtime','steptime'], axis=1)

df2 =df2[1:]
df2 = df2.append(pd.Series([np.nan]*len(df2.columns), index=df2.columns), ignore_index=True)
df3 = df3.iloc[: , 1:]
df_CTR = pd.concat([df3, df2], axis=1)

with pd.ExcelWriter(heatmaps_ctr_RFLF, engine='xlsxwriter') as writer:
      df_CTR.to_excel(writer)
     
# here we do the exact same but for the mutant mice group
swingtime=[]
df = pd.read_excel(alltouches_file_mut)
df = df.iloc[: , 1:]

df = df[df["paw"].str.contains("rh") == False]
df = df[df["paw"].str.contains("lh") == False]
df2=df
for i in range(len(df)-1):
    if (df['paw'].iloc[i] == df['paw'].iloc[i+1]):
        df2 = df2.drop(df.index[i])
    else:
        i=i+1

df2["diff"] = df2["run id"].diff()


    
with pd.ExcelWriter(heatmaps_mut_inbetween, engine='xlsxwriter') as writer:
      df2.to_excel(writer)



df2 = pd.read_excel(heatmaps_mut_inbetween, header = 0, index_col=None)
df2 = df2.iloc[: , 1:]


index_list = df2.index[df2["diff"] != 0].tolist()

result = subtract_from_columns(df2,index_list,value_column) 
for i in range(len(result)-2):
    if result["direction"].iloc[i] == 0:
        result["rung"][i] = (result["rung"][i] - 38)*-1
    swingtime.append(result["begintouches"].iloc[i+2]-result["endtouches"].iloc[i])
swingtime.insert(0,0)
swingtime.append(0)
result["swingtime"] = swingtime
mask = result['diff'] != 0
mask2 = result['swingtime']<=0
result.loc[mask, 'swingtime'] = pd.np.nan
result.loc[mask2, 'swingtime'] = pd.np.nan


merge = [result]
merger=pd.concat(merge,axis=0)
merger = merger.reset_index(drop=True)

indexes = []
indexes2= []
for index, row in merger.iterrows():
  if row['diff'] != 0:
    indexes2.append(index)
    
for index, row in merger.iterrows():
  if row['swingtime'] < 0:
    indexes.append(index)


with pd.ExcelWriter(heatmaps_mut_RFLF, engine='xlsxwriter') as writer:
      merger.to_excel(writer)



df3 = pd.read_excel(heatmaps_mut_RFLF)
steptime = []

for i in range(len(df3)-1):
    steptime.append(df3["touch"].iloc[i]+df3["swingtime"].iloc[i+1])
steptime.insert(0,0) 
df3["steptime"]=steptime
df3 = df3.iloc[: , 1:]

df3.reset_index()
delay = []
duty_cycle=[]
for i in range(len(df3)-1):
    delay.append((df3["begintouches"].iloc[i+1]-df3["begintouches"].iloc[i])/df3["steptime"].iloc[i+1])
    duty_cycle.append((df3["touch"].iloc[i])/df3["steptime"].iloc[i+1])
delay.insert(0,0) 
duty_cycle.insert(0,0) 
df3["delay"]=delay
df3["duty cycle"]=duty_cycle
df2 = df3[['delay', 'swingtime','steptime','duty cycle']]
df3 = df3.drop(['duty cycle','delay', 'swingtime','steptime'], axis=1)

df2 =df2[1:]
df2 = df2.append(pd.Series([np.nan]*len(df2.columns), index=df2.columns), ignore_index=True)
df3 = df3.iloc[: , 1:]
df_MUT = pd.concat([df3, df2], axis=1)

with pd.ExcelWriter(heatmaps_mut_RFLF, engine='xlsxwriter') as writer:
      df_MUT.to_excel(writer)

# filter the data to only get the delay of the RF paw compared to the LH paw not the other way around
# deletes delays smaller then 0 or bigger then 1 these can happen due to a step being unclassified and thus contain nonsense data
# also deletes values without delay this is the last touch a run
df = pd.read_excel(heatmaps_ctr_RFLF, header=0, index_col=None)
df = df[(df.delay >= 0) & (df.delay <= 1) & df["paw"].eq("rf")].dropna(subset=["delay"])
df = df.iloc[:,1:]

dfs = []
for i in range(first_session_number,first_session_number+amount_of_sessions):
    dfs.append(df[df["sessionnr"] == i])

df2 = pd.read_excel(heatmaps_mut_RFLF, header=0, index_col=None)
df2 = df2[(df2.delay >= 0) & (df2.delay <= 1) & df2["paw"].eq("rf")].dropna(subset=["delay"])
df2 = df2.iloc[:,1:]
for i in range(first_session_number,first_session_number+amount_of_sessions):
    dfs.append(df2[df2["sessionnr"] == i])

# Create a Pandas Excel writer using XlsxWriter as the engine
writer = pd.ExcelWriter(HBall_RFLF, engine='xlsxwriter')

# Iterate over the dataframes and write each to a separate sheet
for df in dfs:
    sheet_name = f"{df['subject id'].iloc[0]} - Session {df['sessionnr'].iloc[0]}"
    df.to_excel(writer, sheet_name=sheet_name, index=False)

# Save the Excel file
writer.save()
#%% making the heatmaps 

# loops through the excel sheets, every session has a sheet of data corresponding to this session.
# takes the proper columns of this sheets and plots these values in a 2d histogram from 0-->6.7%

import openpyxl
allvars = []
allsheetnames=[]
file = HBall_RFLF
workbook = openpyxl.load_workbook(file)
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows():
        data.append([cell.value for cell in row])
    df = pd.DataFrame(data)
    x = df[13]
    y = df[10]
    xlist = x.tolist()
    ylist = y.tolist()
    xlist = xlist[2:]
    ylist = ylist[2:]
    array=np.array([xlist,ylist])
    std_dev = np.round(np.std(array),4)
    allsheetnames.append(sheet_name)
    print(sheet_name)
    print(std_dev)
    scaling = len(xlist)*0.067
    xlist = [x * 100 for x in xlist]
    ylist = [y * 100 for y in ylist]
    plt.hist2d(xlist, ylist, bins = ([np.arange(0, 105, 5), np.arange(0, 105, 5)]), vmin=0, vmax=scaling, cmap = 'Greys')
    # plt.axis('off')
    plt.title(sheet_name)
    plt.colorbar()
    plt.savefig(heatmapsRFLF +'\\'+sheet_name+'', format='Tiff',bbox_inches='tight')
    plt.show()
    
#%%
swingtime=[]
df = pd.read_excel(alltouches_file_ctr)
df = df.iloc[: , 1:]

#select the paw pair we want
df = df[df["paw"].str.contains("lf") == False]
df = df[df["paw"].str.contains("lh") == False]
df2=df

#make sure that we dont calculate RF --> RF due to a missing / unclassified touch
for i in range(len(df)-1):
    if (df['paw'].iloc[i] == df['paw'].iloc[i+1]):
        df2 = df2.drop(df.index[i])
    else:
        i=i+1

# determine the diff column to seperate the runs 
df2["diff"] = df2["run id"].diff()


    
with pd.ExcelWriter(heatmaps_ctr_inbetween, engine='xlsxwriter') as writer:
      df2.to_excel(writer)



df2 = pd.read_excel(heatmaps_ctr_inbetween, header = 0, index_col=None)
df2 = df2.iloc[: , 1:]


index_list = df2.index[df2["diff"] != 0].tolist()

# set every run to start at time = 0 by using the diff column
result = subtract_from_columns(df2,index_list,value_column) 
for i in range(len(result)-2):
    if result["direction"].iloc[i] == 0:
        result["rung"][i] = (result["rung"][i] - 38)*-1
    swingtime.append(result["begintouches"].iloc[i+2]-result["endtouches"].iloc[i])
swingtime.insert(0,0)
swingtime.append(0)
result["swingtime"] = swingtime
mask = result['diff'] != 0
mask2 = result['swingtime']<=0
result.loc[mask, 'swingtime'] = pd.np.nan
result.loc[mask2, 'swingtime'] = pd.np.nan


merge = [result]
merger=pd.concat(merge,axis=0)
merger = merger.reset_index(drop=True)

indexes = []
indexes2= []
for index, row in merger.iterrows():
  if row['diff'] != 0:
    indexes2.append(index)
    
for index, row in merger.iterrows():
  if row['swingtime'] < 0:
    indexes.append(index)

with pd.ExcelWriter(heatmaps_ctr_RFRH, engine='xlsxwriter') as writer:
      merger.to_excel(writer)


df3 = pd.read_excel(heatmaps_ctr_RFRH, header = 0, index_col=None)
steptime = []

# add steptime column
for i in range(len(df3)-1):
    steptime.append(df3["touch"].iloc[i]+df3["swingtime"].iloc[i+1])
steptime.insert(0,0) 
df3["steptime"]=steptime
df3 = df3.iloc[: , 1:]

#add delay and duty cycle column
df3.reset_index()
delay = []
duty_cycle=[]
for i in range(len(df3)-1):
    delay.append((df3["begintouches"].iloc[i+1]-df3["begintouches"].iloc[i])/df3["steptime"].iloc[i+1])
    duty_cycle.append((df3["touch"].iloc[i])/df3["steptime"].iloc[i+1])
delay.insert(0,0) 
duty_cycle.insert(0,0) 
df3["delay"]=delay
df3["duty cycle"]=duty_cycle
df2 = df3[['delay', 'swingtime','steptime','duty cycle']]
df3 = df3.drop(['duty cycle','delay', 'swingtime','steptime'], axis=1)

df2 =df2[1:]
df2 = df2.append(pd.Series([np.nan]*len(df2.columns), index=df2.columns), ignore_index=True)
df3 = df3.iloc[: , 1:]
df_CTR = pd.concat([df3, df2], axis=1)

with pd.ExcelWriter(heatmaps_ctr_RFRH, engine='xlsxwriter') as writer:
      df_CTR.to_excel(writer)
     
# here we do the exact same but for the mutant mice group
swingtime=[]
df = pd.read_excel(alltouches_file_mut)
df = df.iloc[: , 1:]

df = df[df["paw"].str.contains("lf") == False]
df = df[df["paw"].str.contains("lh") == False]
df2=df
for i in range(len(df)-1):
    if (df['paw'].iloc[i] == df['paw'].iloc[i+1]):
        df2 = df2.drop(df.index[i])
    else:
        i=i+1

df2["diff"] = df2["run id"].diff()


    
with pd.ExcelWriter(heatmaps_mut_inbetween, engine='xlsxwriter') as writer:
      df2.to_excel(writer)



df2 = pd.read_excel(heatmaps_mut_inbetween, header = 0, index_col=None)
df2 = df2.iloc[: , 1:]


index_list = df2.index[df2["diff"] != 0].tolist()

result = subtract_from_columns(df2,index_list,value_column) 
for i in range(len(result)-2):
    if result["direction"].iloc[i] == 0:
        result["rung"][i] = (result["rung"][i] - 38)*-1
    swingtime.append(result["begintouches"].iloc[i+2]-result["endtouches"].iloc[i])
swingtime.insert(0,0)
swingtime.append(0)
result["swingtime"] = swingtime
mask = result['diff'] != 0
mask2 = result['swingtime']<=0
result.loc[mask, 'swingtime'] = pd.np.nan
result.loc[mask2, 'swingtime'] = pd.np.nan


merge = [result]
merger=pd.concat(merge,axis=0)
merger = merger.reset_index(drop=True)

indexes = []
indexes2= []
for index, row in merger.iterrows():
  if row['diff'] != 0:
    indexes2.append(index)
    
for index, row in merger.iterrows():
  if row['swingtime'] < 0:
    indexes.append(index)


with pd.ExcelWriter(heatmaps_mut_RFRH, engine='xlsxwriter') as writer:
      merger.to_excel(writer)



df3 = pd.read_excel(heatmaps_mut_RFRH)
steptime = []

for i in range(len(df3)-1):
    steptime.append(df3["touch"].iloc[i]+df3["swingtime"].iloc[i+1])
steptime.insert(0,0) 
df3["steptime"]=steptime
df3 = df3.iloc[: , 1:]

df3.reset_index()
delay = []
duty_cycle=[]
for i in range(len(df3)-1):
    delay.append((df3["begintouches"].iloc[i+1]-df3["begintouches"].iloc[i])/df3["steptime"].iloc[i+1])
    duty_cycle.append((df3["touch"].iloc[i])/df3["steptime"].iloc[i+1])
delay.insert(0,0) 
duty_cycle.insert(0,0) 
df3["delay"]=delay
df3["duty cycle"]=duty_cycle
df2 = df3[['delay', 'swingtime','steptime','duty cycle']]
df3 = df3.drop(['duty cycle','delay', 'swingtime','steptime'], axis=1)

df2 =df2[1:]
df2 = df2.append(pd.Series([np.nan]*len(df2.columns), index=df2.columns), ignore_index=True)
df3 = df3.iloc[: , 1:]
df_MUT = pd.concat([df3, df2], axis=1)

with pd.ExcelWriter(heatmaps_mut_RFRH, engine='xlsxwriter') as writer:
      df_MUT.to_excel(writer)

# filter the data to only get the delay of the RF paw compared to the LH paw not the other way around
# deletes delays smaller then 0 or bigger then 1 these can happen due to a step being unclassified and thus contain nonsense data
# also deletes values without delay this is the last touch a run
df = pd.read_excel(heatmaps_ctr_RFRH, header=0, index_col=None)
df = df[(df.delay >= 0) & (df.delay <= 1) & df["paw"].eq("rf")].dropna(subset=["delay"])
df = df.iloc[:,1:]

dfs = []
for i in range(first_session_number,first_session_number+amount_of_sessions):
    dfs.append(df[df["sessionnr"] == i])

df2 = pd.read_excel(heatmaps_mut_RFRH, header=0, index_col=None)
df2 = df2[(df2.delay >= 0) & (df2.delay <= 1) & df2["paw"].eq("rf")].dropna(subset=["delay"])
df2 = df2.iloc[:,1:]
for i in range(first_session_number,first_session_number+amount_of_sessions):
    dfs.append(df2[df2["sessionnr"] == i])

# Create a Pandas Excel writer using XlsxWriter as the engine
writer = pd.ExcelWriter(HBall_RFRH,engine='xlsxwriter')

# Iterate over the dataframes and write each to a separate sheet
for df in dfs:
    sheet_name = f"{df['subject id'].iloc[0]} - Session {df['sessionnr'].iloc[0]}"
    df.to_excel(writer, sheet_name=sheet_name, index=False)

# Save the Excel file
writer.save()
#%% making the heatmaps 

# loops through the excel sheets, every session has a sheet of data corresponding to this session.
# takes the proper columns of this sheets and plots these values in a 2d histogram from 0-->6.7%

import openpyxl
allvars = []
allsheetnames=[]
file = HBall_RFRH
workbook = openpyxl.load_workbook(file)
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows():
        data.append([cell.value for cell in row])
    df = pd.DataFrame(data)
    x = df[13]
    y = df[10]
    xlist = x.tolist()
    ylist = y.tolist()
    xlist = xlist[2:]
    ylist = ylist[2:]
    array=np.array([xlist,ylist])
    std_dev = np.round(np.std(array),4)
    allsheetnames.append(sheet_name)
    print(sheet_name)
    print(std_dev)
    scaling = len(xlist)*0.067
    xlist = [x * 100 for x in xlist]
    ylist = [y * 100 for y in ylist]
    plt.hist2d(xlist, ylist, bins = ([np.arange(0, 105, 5), np.arange(0, 105, 5)]), vmin=0, vmax=scaling, cmap = 'Greys')
    # plt.axis('off')
    plt.title(sheet_name)
    plt.colorbar()
    plt.savefig(heatmapsRFRH +'\\'+sheet_name+'', format='Tiff',bbox_inches='tight')
    plt.show()
#%% circularity/ excentricity 
df = pd.read_excel(lag_file_mut)
df2 = pd.read_excel(lag_file_ctr)

df = df[df["pair"]=="rf/lh"]
df2 = df2[df2["pair"]=="rf/lh"]
# change the amount of bins from 20 to 10.
df["lag"] = df["lag"].div(2)
df2["lag"]=df2["lag"].div(2)
df["lag"] = round(df["lag"],1)
df2["lag"] = round(df2["lag"],1)

data_frames = {}
data_frames2={}
for i in range(first_session_number,first_session_number+amount_of_sessions):
    data_frames[f"ko_{i}"] = df[df["ses nr"] == i]
    data_frames2[f"wt_{i}"] = df2[df2["ses nr"] == i]

    
    data_frames[f"ko_{i}"] = data_frames[f"ko_{i}"].groupby(["subject id"])["lag"].value_counts()
    data_frames2[f"wt_{i}"] = data_frames2[f"wt_{i}"].groupby(["subject id"])["lag"].value_counts()
dfmut = pd.concat(data_frames,axis=1)
dfctr = pd.concat(data_frames2,axis=1)


with pd.ExcelWriter(lag_10bins, engine='xlsxwriter') as writer:
      dfmut.to_excel(writer, sheet_name = 'KO')
      dfctr.to_excel(writer, sheet_name = 'WT')


import shapely.geometry as geometry
def pol2cart(theta, rho):
    x = rho * np.cos(theta)
    y = rho * np.sin(theta)
    return x, y


from shapely.geometry import Point


# Define the input and output file paths
input_file = lag_10bins
output_file = circularity_and_centroidsWT

# Define a list of subject IDs to process

# Create an Excel writer object to write the results to the output file
writer = pd.ExcelWriter(output_file)

for subject_id in subject_ids_WT:

    # Load the data for the current subject ID
    df = pd.read_excel(input_file, sheet_name='WT')
    # step to solve a binning artifiact causing the 0 and 1 bin to be added to creat a similair size bin as the others.
    for i in range(len(df)):
        # check if the subject id is NaN
        if pd.isna(df.loc[i, "subject id"]):
            # replace NaN with the subject id from the row above
            df.loc[i, "subject id"] = df.loc[i-1, "subject id"]
    df= df.fillna(0)

    df_subject = df[df["subject id"] == subject_id]
    rows_to_sum = df_subject.loc[df_subject['lag'].isin([0, 1])]
     
    row_sum = rows_to_sum.sum()
    df_subject.loc[df_subject.index.min()] = row_sum
    df_subject = df_subject.drop(rows_to_sum.index)
    df_subject.loc[df_subject['lag'] == 1, 'lag'] = 0
    df_subject = df_subject.sort_values(by=["lag"])

    circularlist = []
    centroidlist = []

    for i in range(first_session_number,first_session_number+amount_of_sessions):

         data = df_subject[f'wt_{i}'].tolist()
         data.append(data[0])
         label_loc = np.linspace(start=0, stop=2 * np.pi, num=len(data))
         x2, y2 = pol2cart(label_loc, data)
         polygon = geometry.Polygon(zip(x2, y2))
         area = polygon.area
         x, y = pol2cart(label_loc, data)
         polygon = geometry.Polygon(zip(x, y))
         perimeter = polygon.length
         if perimeter ==0:
             continue
         circularity = 4 * np.pi * area / perimeter ** 2
         circularity = round(circularity, 3)
         circularlist.append(circularity)
         COG = polygon.centroid
         centroid_distance = Point(0, 0).distance(COG)
         centroidlist.append(centroid_distance)

         plt.figure(figsize=(8, 8))
         plt.subplot(polar=True)
         plt.plot(label_loc, data, label='data')
         plt.legend()
         plt.show()

     # Create a DataFrame to store the results
    results_df = pd.DataFrame({"Circularity": circularlist, "Centroid Distance": centroidlist})

     # Write the results to a sheet with the current subject ID
    results_df.to_excel(writer, sheet_name=f"Subject {subject_id}", index=False)

 # Save and close the Excel file
writer.save()

input_file = lag_10bins
output_file = circularity_and_centroidsKO

# Define a list of subject IDs to process

# Create an Excel writer object to write the results to the output file
writer = pd.ExcelWriter(output_file)

for subject_id in subject_ids_KO:

    # Load the data for the current subject ID
    df = pd.read_excel(input_file, sheet_name='KO')
    for i in range(len(df)):
        # check if the subject id is NaN
        if pd.isna(df.loc[i, "subject id"]):
            # replace NaN with the subject id from the row above
            df.loc[i, "subject id"] = df.loc[i-1, "subject id"]
    df= df.fillna(0)
    df_subject = df[df["subject id"] == subject_id]
    rows_to_sum = df_subject.loc[df_subject['lag'].isin([0, 1])]
    
    row_sum = rows_to_sum.sum()
    df_subject.loc[df_subject.index.min()] = row_sum

    df_subject = df_subject.drop(rows_to_sum.index)
    df_subject.loc[df_subject['lag'] == 1, 'lag'] = 0
    df_subject = df_subject.sort_values(by=["lag"])

    circularlist = []
    centroidlist = []

    for i in range(first_session_number,first_session_number+amount_of_sessions):

        data = df_subject[f'ko_{i}'].tolist()
        data.append(data[0])
        label_loc = np.linspace(start=0, stop=2 * np.pi, num=len(data))
        x2, y2 = pol2cart(label_loc, data)
        polygon = geometry.Polygon(zip(x2, y2))
        area = polygon.area
        x, y = pol2cart(label_loc, data)
        polygon = geometry.Polygon(zip(x, y))
        perimeter = polygon.length
        if perimeter ==0:
            continue
        circularity = 4 * np.pi * area / perimeter ** 2
        circularity = round(circularity, 3)
        circularlist.append(circularity)
        COG = polygon.centroid
        centroid_distance = Point(0, 0).distance(COG)
        centroidlist.append(centroid_distance)

        plt.figure(figsize=(8, 8))
        plt.subplot(polar=True)
        plt.plot(label_loc, data, label='data')
        plt.legend()
        plt.show()

    # Create a DataFrame to store the results
    results_df = pd.DataFrame({"Circularity": circularlist, "Centroid Distance": centroidlist})

    # Write the results to a sheet with the current subject ID
    results_df.to_excel(writer, sheet_name=f"Subject {subject_id}", index=False)

# Save and close the Excel file
writer.save()
#%% support circles 
ercc1_mut = pd.read_excel(parameter_file_mut)
ercc1_ctr = pd.read_excel(parameter_file_ctr)
df = pd.DataFrame(ercc1_ctr)
df2 = pd.DataFrame(ercc1_mut)



ctr = df.groupby(["session nr","subject id"]).median()
ctr = ctr.drop(["direction", "timerun", "number of steps", "even steps", "odd steps","run id", "step 6", "step 2", "step 4", "swing phase lh", "swing phase rh","swing phase lf", "stance phase rf", "stance phase lf", "stance phase rh","stance phase lh", "lag lf/rh", "lag rf/lh", "swing phase rf", "session id","high rungs front", "high rungs hind", "low rungs front", "low rungs hind"], axis=1)



mut = df2.groupby(["session nr","subject id"]).median()
mut = mut.drop(["direction", "timerun", "number of steps", "even steps", "odd steps","run id", "step 6", "step 2", "step 4", "swing phase lh", "swing phase rh",                "swing phase lf", "stance phase rf", "stance phase lf", "stance phase rh",                "stance phase lh", "lag lf/rh", "lag rf/lh", "swing phase rf", "session id",                "high rungs front", "high rungs hind", "low rungs front", "low rungs hind"], axis=1)



with pd.ExcelWriter(supports, engine='xlsxwriter') as writer:
    mut.to_excel(writer, sheet_name='mut')
    ctr.to_excel(writer, sheet_name='ctr')
#%%
