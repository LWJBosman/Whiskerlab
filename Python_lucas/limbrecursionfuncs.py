import xlrd
import xlsxwriter
import openpyxl
from bokeh.plotting import figure, output_file, show, save
from openpyxl.styles import Font
import os.path

from Python.move import Move


def getsteppatternfromfile(trial_id):
    filepath = "Trialstest\\Trial_%s.xlsx" % trial_id
    file_exists = True
    if os.path.isfile(filepath):
        print("File exist")
        file_exists = True
    else:
        print("File not exist")
        file_exists = False
        return None

    if file_exists:
        workbook = xlrd.open_workbook(filepath)
        worksheet = workbook.sheet_by_name("Moves")
        nrofsteps = len(worksheet.col_values(0))
        sessionnr = worksheet.cell_value(8, 21)
        moves = []
        for row in range(1, nrofsteps):
            move_id = worksheet.cell_value(row, 0)
            if move_id == "":
                print("Probably no moves found for trial or < 10")
                break
            else:
                move_id = int(move_id)
            limb = int(worksheet.cell_value(row, 1))
            rung_start = int(worksheet.cell_value(row, 2))
            rung_end = int(worksheet.cell_value(row, 3))
            t_start = int(worksheet.cell_value(row, 5))
            t_fly_start = int(worksheet.cell_value(row, 6))
            t_end = int(worksheet.cell_value(row, 7))
            touch_duration = worksheet.cell_value(row, 9)
            if touch_duration == "":
                touch_duration = 0
            else:
                touch_duration = int(touch_duration)
            touchside = worksheet.cell_value(row, 10)
            if touchside == "":
                pass
            else:
                touch_duration = int(touch_duration)
            p_noise = int(worksheet.cell_value(row, 12))
            p_scores = worksheet.cell_value(row, 13)
            likeliness = int(worksheet.cell_value(row, 14))
            moves.append([move_id, Move(limb=limb, rung_start=rung_start, rung_end=rung_end,
                                        touchduration=touch_duration, t_start=t_start, t_fly_start=t_fly_start,
                                        t_end=t_end, p_noise=p_noise, p_scores=p_scores, likeliness=likeliness)])
        return moves, filepath, sessionnr

class Limbrecursionfilesfuncs(object):
    def __init__(self, logger):
        self.logger = logger
        pass

    def plot_steppattern(self, trial_id):
        filepath = 'Trials\\Trial_%s.xlsx' % trial_id
        workbook = xlrd.open_workbook(filepath)
        worksheet = workbook.sheet_by_name("Moves")

        p = figure(x_axis_label='time (ms)',
                   y_axis_label='rung',
                   x_range=(0, 5000),
                   y_range=(0, 38),
                   plot_width=800)
        p.xaxis.axis_label_text_font_size = "20pt"
        p.yaxis.axis_label_text_font_size = "20pt"

        colors = ['red', 'green', 'purple', 'orange']
        limbs = ['LF limb', 'RF limb', 'LH limb', 'RH limb']
        nrofsteps = len(worksheet.col_values(0))
        nrofstepsinplot = 0

        for row in range(1, nrofsteps):
            limb = worksheet.cell_value(row, 1)
            if not limb:
                print("all touches processed")
                break
            else:
                limb = int(limb)
            rung = int(worksheet.cell_value(row, 2))
            touch_start = int(worksheet.cell_value(row, 5))
            touch_end = int(worksheet.cell_value(row, 6))
            touch_side = worksheet.cell_value(row, 10)
            if touch_side == "":
                continue
            else:
                touch_side = int(touch_side)
            if (rung % 2 == 0 and touch_side == 0) or (rung % 2 == 1 and touch_side == 1):
                p.circle_cross(touch_start, rung, line_color=colors[limb - 1], size=8, legend=limbs[limb - 1])
            else:
                p.circle(touch_start, rung, line_color=colors[limb - 1], fill_color=colors[limb - 1])
            p.line([touch_start, touch_end], [rung, rung], line_color=colors[limb - 1], legend=limbs[limb - 1])
            nrofstepsinplot += 1

        p.title.text = "walking pattern mouse | trial_id = %s | steps = %s " % (trial_id, nrofstepsinplot)
        output_file("Trials\\trial %s_limbdetect.html" % trial_id)
        save(p)

    def score_pattern(self, trial_id):
        print("Attempting to score trial %s" % trial_id)
        moves, filepath, sessionnr = getsteppatternfromfile(trial_id)
        if moves:
            print("Scoring trial %s" % trial_id)
            limb1 = 0
            limb2 = 0
            limb3 = 0
            limb4 = 0
            if len(moves) > 0:
                for m in moves:
                    if m[1].limb == 1:
                        limb1 += 1
                    if m[1].limb == 2:
                        limb2 += 1
                    if m[1].limb == 3:
                        limb3 += 1
                    if m[1].limb == 4:
                        limb4 += 1
                avg = sum([limb1, limb2, limb3, limb4]) / 4
                differences = [sum([limb2, limb3, limb4]) - 3 * limb1,
                               sum([limb1, limb3, limb4]) - 3 * limb2,
                               sum([limb1, limb2, limb4]) - 3 * limb3,
                               sum([limb1, limb2, limb3]) - 3 * limb4]
                max_difference = [differences.index(max(differences)) + 1,
                                  max(differences),
                                  max(differences) / avg,  # expected to be most useful
                                  differences.index(min(differences)) + 1,
                                  min(differences),
                                  min(differences) / avg]
            else:
                max_difference = [0, 100, 100, 0, -100, -100]

            workbook = openpyxl.load_workbook(filepath)
            worksheet = workbook.get_sheet_by_name('Moves')

            baserow = 12
            score_names = ['min limb', 'sum difference', 'sum/avg', 'max limb', 'sum differences', 'sum/avg']
            limbnames = ['limb 1', 'limb 2', 'limb 3', 'limb 4']
            limbs = [limb1, limb2, limb3, limb4]
            for i in range(len(limbnames)):
                worksheet.cell(baserow, i + 18).value = limbnames[i]
                worksheet.cell(baserow, i + 18).font = Font(bold=True)
                worksheet.cell(baserow + 1, i + 18).value = limbs[i]

            for i in range(len(score_names)):
                worksheet.cell(baserow + 2, i + 18).value = score_names[i]
                worksheet.cell(baserow + 2, i + 18).font = Font(bold=True)
                worksheet.cell(baserow + 3, i + 18).value = max_difference[i]

            workbook.save(filepath)
            workbook.close()

            return max_difference[2], max_difference[5], sessionnr

    def labeltrials(self, database):
        controlercc16 = [189, 194, 195, 197, 199, 200, 202, 204]  # exp. no 16
        mutantercc16 = [190, 191, 193, 196, 198, 201, 203, 205]
        src_filepath = 'Trialstest\\Trialscores.xlsx'
        to_filepath = 'Trialstest\\Trialslabeled.xlsx'
        src_workbook = openpyxl.load_workbook(src_filepath)
        src_sheet = src_workbook.get_sheet_by_name('Scores')
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet('Trials')
        row = 1
        sessions = 19
        for i in range(1, sessions + 1):
            worksheet.cell(row, 2 * i - 1).value = "session %s" % i
            worksheet.cell(row, 2 * i - 1).font = Font(bold=True)
            worksheet.cell(row + 1, 2 * i - 1).value = "control"
            worksheet.cell(row + 1, 2 * i - 1).font = Font(bold=True)
            worksheet.cell(row + 1, 2 * i).value = "mutant"
            worksheet.cell(row + 1, 2 * i).font = Font(bold=True)

        currentsession = 1
        column = 1
        controlrow = 3
        mutantrow = 3
        breakafterthis=False
        while True:
            row += 1
            trial_id = src_sheet.cell(row, 1).value
            print(trial_id)
            yes = src_sheet.cell(row, 5).value
            session = src_sheet.cell(row, 4).value
            if trial_id == 64432:
                breakafterthis = True
            if not yes:
                continue
            if session == '0':
                continue
            if int(session) > currentsession:
                currentsession = int(session)
                column += 2
                controlrow = 3
                mutantrow = 3
            query = database.query_builder([('DISTINCT(subject', 'id)'), ('subject', 'identifier')],
                                           jointables=['subject', 'session', 'trial'],
                                           var_eq1=['trial', 'id', int(trial_id), '='])
            subjectdata = database.query(query)
            if subjectdata[0][0] in controlercc16:
                worksheet.cell(controlrow, column).value = trial_id
                controlrow += 1
            if subjectdata[0][0] in mutantercc16:
                worksheet.cell(mutantrow, column + 1).value = trial_id
                mutantrow += 1
            workbook.save(to_filepath)
            if breakafterthis:
                break

        workbook.close()
